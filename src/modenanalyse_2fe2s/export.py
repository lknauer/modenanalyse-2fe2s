# Copyright (C) 2026 Lukas Knauer, AG Schünemann, RPTU Kaiserslautern-Landau
# SPDX-License-Identifier: GPL-3.0-or-later
# Teil von modenanalyse_2fe2s — see LICENSE in the Wurzelverzeichnis.

# -*- coding: utf-8 -*-
"""
export.py
======================
Excel-Export the Analyseergebnisse in sechs separate Excel fileen
plus Embedding-PNGs.

output files
--------------
``_analysis.xlsx``
    Core analysis: mode_analysis, groups, Fe-ligands, His_HN, kernel scores,
    distances, SCSD, reorganization energies, B-factors, info.
``_analysis_SSE.xlsx``
    Secondary-structure amplitudes (only if SSE elements present in the PDB).
``_analysis_Embeddings.xlsx``
    UMAP coordinates, cluster sheets, C-alpha amplitudes.
``_analysis_interp{step}.xlsx``
    Interpolated core analysis on a uniform grid
    (step size = cfg.interp_step). Symmetric boundary treatment:
    context modes left (context_results_left) and right (context_results).
``_analysis_SSE_interp{step}.xlsx``
    Secondary-structure amplitudes on a uniform grid.

Public functions
-----------------------
export_main_excel
    Writes ``_analysis.xlsx``.
export_sse_excel
    Writes ``_analysis_SSE.xlsx``.
export_embedding_excel
    Writes ``_analysis_Embeddings.xlsx``.
export_interpolated_excel
    Writes ``_analysis_interp.xlsx``.
export_embedding_plots
    Writes Embedding-PNGs (Frequenz- and Modentyp-Faerbung).

Bugfixes (gegenvia Vorversion)
---------------------------------
B4  ``*_Cluster``-Sheets: Feature-Spalten are jetzt befuellt.
"""
from __future__ import annotations
import os, time
from typing import Dict, List, Optional, Tuple
import numpy as np
from .config   import Config
from .logio       import RunLog
from .geometry import CoordInfo

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils  import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

# ===========================================================================
# ExportPayload  (Hardening #7)
# ===========================================================================

from dataclasses import dataclass, field as _field


@dataclass
class ExportPayload:
    """Buendelt alle Analyseergebnisse for the Export.

    Wird von modenanalyse.main() befuellt and an export_all() uebergeben.
    Ersetzt the bisherigen losen Parameter listn the einzelnen
    export_*-functions. Zukuenftige Erweiterungen are hier als
    optionale Felder ergaenzt -- the Signatur von export_all() bleibt stabil.

    Parameters
    ----------
    results : list of dict
        Modenanalyse-Ergebnisse (sortiert after Frequenz).
    coord_info : CoordInfo
        Koordinations-Information.
    dist_ref : dict
        Gleichgewichtsabstaende.
    logname : str
        Basisname the Gaussian-Logdatei.
    cfg : Config
        Laufkonfiguration.
    runlog : RunLog
        Protokoll-Objekt.
    """

    # Pflichtfelder
    results:       list
    coord_info:    object
    dist_ref:      dict
    logname:       str
    cfg:           object
    runlog:        object

    # Optionale Felder
    cluster_info:         object  = None
    b_factors:            object  = None
    atoms:                object  = None
    context_results:      list    = _field(default_factory=list)
    context_results_left: list    = _field(default_factory=list)

    # Embeddings
    embedding_coords:   dict    = _field(default_factory=dict)
    embed_feat_matrix:  object  = None
    embed_feat_names:   list    = _field(default_factory=list)
    cluster_data:       dict    = _field(default_factory=dict)
    ca_data:            object  = None
    sse_umap_data:       object  = None
    ca_umap_data:       object  = None        # NEW in v1.0.3: Ca-UMAP tuple from compute_ca_umap_cluster


def export_all(payload):
    """Writes alle sechs Excel-output fileen and the Embedding-PNGs.

    Einziger stabiler Einstiegspunkt for the Export.  Alle
    Teilergebnisse are over payload uebergeben, not als
    einzelne lose Parameter.

    Parameters
    ----------
    payload : ExportPayload
        Vollstaendig befuelltes Export-Nutzlast-Objekt.

    Notes
    -----
    Hardening #7: Diese Funktion ersetzt the direkten Aufrufe the sechs
    einzelnen export_*-functions from modenanalyse.main().
    Zukuenftige Erweiterungen are als Felder in ExportPayload
    ergaenzt, without the Signatur dieser Funktion to aendern.
    """
    p   = payload
    cfg = p.cfg
    step = cfg.interp_step

    export_main_excel(
        p.results, p.coord_info, p.dist_ref,
        p.logname, cfg.outname("_analysis.xlsx"),
        cfg, p.runlog,
        cluster_info=p.cluster_info,
        b_factors=p.b_factors,
        atoms=p.atoms,
    )

    if any(r.get("sse") for r in p.results):
        export_sse_excel(
            p.results, cfg.outname("_analysis_SSE.xlsx"),
            cfg, p.runlog, sse_umap_data=p.sse_umap_data,
        )
    elif cfg.analyze_sse:
        p.runlog.warn(
            "SSE-Export: no 'sse' key in results found. "
            "_analysis_SSE.xlsx will not be created. "
            "Cause: SSE analysis failed or sse_center_map empty.")

    if p.embedding_coords or p.ca_data:
        export_embedding_excel(
            p.results, p.embedding_coords,
            p.embed_feat_matrix, p.embed_feat_names,
            p.cluster_data, p.ca_data,
            cfg.outname("_analysis_Embeddings.xlsx"),
            cfg, p.runlog,
            ca_umap_data=p.ca_umap_data,
        )

    try:
        export_interpolated_excel(
            p.results, p.coord_info,
            cfg.outname(f"_analysis_interp{step:.2f}.xlsx"),
            cfg, p.runlog,
            context_results=p.context_results,
            context_results_left=p.context_results_left,
        )
    except Exception as exc:
        p.runlog.warn(f"interpolation Excel failed: {exc}")

    try:
        export_sse_interp_excel(
            p.results,
            cfg.outname(f"_analysis_SSE_interp{step:.2f}.xlsx"),
            cfg, p.runlog,
            context_results=p.context_results,
            context_results_left=p.context_results_left,
        )
    except Exception as exc:
        p.runlog.warn(f"SSE-interpolation Excel failed: {exc}")

    # Hardening v3.1: Embedding-PNGs are off by default. Set
    # cfg.export_embedding_plots = True to enable them. They cost time
    # and produce questionable images when HDBSCAN finds nothing.
    # v1.0.3: now also renders SSE-UMAP, Ca-UMAP and Ca-amplitude heatmap
    # PNGs if the corresponding data has been computed.
    if getattr(cfg, "export_embedding_plots", False):
        export_embedding_plots(p.embedding_coords, p.results,
                                cfg.outname, p.runlog,
                                sse_umap_data=p.sse_umap_data,
                                ca_umap_data=p.ca_umap_data,
                                ca_data=p.ca_data)


_brd = None
def _bd():
    """Returns a duennen grauen Rahmen (Border) for Excel-Zellen  (Singleton)."""
    global _brd
    if _brd is None and _HAS_OPENPYXL:
        s = Side(style="thin", color="CCCCCC")
        _brd = Border(left=s, right=s, top=s, bottom=s)
    return _brd

def _hc(ws, row, col, text, width, fg="1F4E79"):
    """Writes a formatierte Kopfzeilen-Zelle (weiss, bold, farbiger Hintergrund).

    Parameters
    ----------
    ws : Worksheet
        Ziel-Sheet.
    row, col : int
        Zeilen- and Spaltenindex (1-basiert).
    text : str
        Zelleninhalt.
    width : int or float
        Spaltenbreite in Excel-Einheiten (0 = not setzen).
    fg : str, optional
        Hintergrundfarbe als 6-stelliger Hex-Code. Standard: Dunkelblau.
    """
    c = ws.cell(row, col, text)
    c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    c.fill      = PatternFill("solid", fgColor=fg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    b = _bd()
    if b: c.border = b
    if width: ws.column_dimensions[get_column_letter(col)].width = width

def _dc(ws, row, col, v, fill=None):
    """Writes einen Datenwert rechtbuendig in a Excel-Zelle.

    Rundet float-values on 6 Dezimalstellen; ersetzt nicht-endliche values through None.
    """
    c = ws.cell(row, col)
    c.value = (round(v,6) if isinstance(v,float) and np.isfinite(v) else
               (v if not isinstance(v,float) else None))
    c.font      = Font(name="Arial", size=9)
    c.alignment = Alignment(horizontal="right", vertical="center")
    b = _bd()
    if b: c.border = b
    if fill: c.fill = PatternFill("solid", fgColor=fill)

def _sc(ws, row, col, v, fill=None):
    """Writes einen Unsicherheitswert als Zahl (kursiv, grau) in a Excel-Zelle.

    Der value is float geschrieben (kein "±"-Praefix), so that Excel numerisch
    sortieren and filtern kann. Der zugehoerige Spalten-Header contributes the ±-Zeichen.
    """
    c = ws.cell(row, col)
    # Als Zahl schreiben: erpossiblet numerisches Sortieren in Excel
    c.value = (round(v, 8) if isinstance(v, float) and np.isfinite(v) else None)
    c.font  = Font(name="Arial", size=8, italic=True, color="888888")
    c.alignment = Alignment(horizontal="right", vertical="center")
    b = _bd()
    if b: c.border = b
    if fill: c.fill = PatternFill("solid", fgColor=fill)

from .core import SCORE_KEYS as _SCORE_KEYS

_MTYPE_FILL = {"Out-of-plane":"FFF2CC","In-plane":"E2EFDA"}
_HIS_RES    = {"HIS","HIE","HID","HIP","HSD","HSE","HSP"}

# v3.5: Farbcodierung after Signifikanz (universelle Konvention)
#   trivial      → hellrot     (FFE0E0)  |X|/sigma <= low
#   signifikant  → hellgelb    (FFF8E0)  low < |X|/sigma <= high
#   hoch         → hellgruen   (E0F0E0)  |X|/sigma > high
# If the Mode type-Zeilenfaerbung gesetzt ist, hat the Zellen-Faerbung
# Vorrang for the Signifikanz-Anzeige (Zeile bleibt over unfaerbe
# Spalten gefaerbt).
_SIG_FILL = {
    "trivial":     "FFE0E0",
    "significant": "FFF8E0",
    "high":        "E0F0E0",
}


def _sig_fill(value: float, sigma: float,
              thr_low: float = 1.0, thr_high: float = 3.0):
    """Farbcode-Hex for Signifikanz a valuess; fallback Mode type-Fill.

    Verwendet the universelle Klassifikation from core.classify_significance.
    returns None zurueck if value or Sigma not endlich (Zelle unfaerbe).
    """
    if value is None or sigma is None:
        return None
    if not isinstance(value, (int, float)) or not isinstance(sigma, (int, float)):
        return None
    if not np.isfinite(value) or not np.isfinite(sigma):
        return None
    from .core import classify_significance
    klass = classify_significance(float(value), float(sigma), thr_low, thr_high)
    return _SIG_FILL.get(klass)


def _dc_sig(ws, row, col, v, sigma, row_fill=None,
            thr_low: float = 1.0, thr_high: float = 3.0):
    """Writes einen Datenwert with Signifikanz-basierter Zellen-Faerbung.

    If Zellen-Faerbung from Signifikanz gives, ueberschreibt sie
    the Mode type-Zeilenfaerbung for diese a Zelle. Das is gewuenscht:
    so sieht the Anwender sofort, welche values in a OOP- oder
    INP-Mode signifikant are and welche in the noise verschwinden.
    """
    fill = _sig_fill(v, sigma, thr_low, thr_high) or row_fill
    _dc(ws, row, col, v, fill=fill)


def _save(wb, outfile, runlog):
    """Saves the Arbeitsmappe, gibt the filenamen from and registriert sie in the RunLog."""
    wb.save(outfile)
    print(f"  -> {os.path.basename(outfile)}")
    runlog.add_output(outfile)


# ===========================================================================
# 1) HAUPT-EXCEL
# ===========================================================================

def export_main_excel(
        results:      List[Dict],
        coord_info:   CoordInfo,
        dist_ref:     Dict,
        logname:      str,
        outfile:      str,
        cfg:          Config,
        runlog:       RunLog,
        cluster_info: Optional[List]       = None,
        b_factors:    Optional[np.ndarray] = None,
        atoms:        Optional[List]       = None,
):
    """Writes the Kern-Analyse in ``_analysis.xlsx``.

    Parameters
    ----------
    results : list of dict
        Modenanalyse-Ergebnisse.
    coord_info : CoordInfo
        Koordinations-Information.
    dist_ref : dict
        Gleichgewichtsabstaende from ``compute_dist_ref``.
    logname : str
        Basisname the Quelldatei (fuer Info-Sheet).
    outfile : str
        Zieldatei.
    cfg : Config
        Konfiguration.
    runlog : RunLog
        Fuer Warnmeldungen and filepfad-Registrierung.
    cluster_info : list of tuple, optional
        Cluster-Atom-Informationen for the Info-Sheet.
    b_factors : ndarray of shape (n_heavy,), optional
        Atomare Debye-Waller-Faktoren in A^2 (aus normal modenanalyse).
    atoms : list of dict, optional
        Gaussian-atom list (fuer B-Faktor-Sheet).

    Notes
    -----
    Sheets: Modenanalyse | Gruppen_OOP/INP/Winkel/Tors | Fe_ligands |
    His_HN (nur if His protoniert) | Kern_Scores | Abstaende |
    SCSD (wenn computed) | B_Faktoren (wenn computed) | Info.
    """
    if not _HAS_OPENPYXL:
        runlog.warn("openpyxl missing."); return

    wb = Workbook(); wb.remove(wb.active)
    E  = cfg.show_errors_in_excel
    gn = list(coord_info.group_map.keys())
    his_prot = [l for l in coord_info.ligands
                if l.res_name.upper() in _HIS_RES and l.lig_element=="N"
                and l.his_protonated]

    if getattr(cfg, "analysis_compact", True):
        _ws_modenanalyse(wb, results, E, cfg)
    if getattr(cfg, "analysis_full", False):
        _ws_modenanalyse_voll(wb, results, cfg)
    if gn: _ws_gruppen(wb, results, gn, E)
    # Fe-S (Cys) and Fe-N (His) in getrennten Sheets with Abschnittsheadern
    _ws_fe_bindung(wb, results, coord_info, "S", E, cfg=cfg)
    _ws_fe_bindung(wb, results, coord_info, "N", E, cfg=cfg)
    if cfg.include_hn_vibration and his_prot:
        if any(r.get("his_hn") for r in results):
            _ws_his_hn(wb, results, his_prot, E)
    _ws_kern_scores(wb, results, E)
    _ws_abstaende(wb, dist_ref)
    scsd_r = [r for r in results if r.get("scsd")]
    if scsd_r: _ws_scsd(wb, results, scsd_r)
    # v3.7: Marcus-Hush-Reorganisations-Sheets (ohne Klassifikation,
    # without Schwellen, without Filter — only reine Modulations-Daten und
    # System-Aggregate. Banden-Identifikation occurs extern in Origin
    # through the Anwender).
    if any(r.get("reorg_per_mode") for r in results):
        # Pro-Mode-Diagnose: dr_X and lambda_X for alle channels
        _ws_reorganisationsenergie_v37(wb, results)
        # System-Aggregate: Total-Reorg per Kanal (eine Zahl per Kanal)
        # and kumulative Lambda(omega), Modulations-Spektren M_X(omega).
        # These Sheets are befuellt, if the runner sie als
        # additionallye Felder durchreicht (see runner.py).
        if any(r.get("_v37_aggregates") for r in [results[0]] if results):
            _ws_reorg_total_v37(wb, results[0]["_v37_aggregates"])
            _ws_reorg_pro_bindung_v37(wb, results[0]["_v37_aggregates"])
            _ws_modulations_spektren_v37(wb, results[0]["_v37_aggregates"])
            _ws_lambda_kumulativ_v37(wb, results[0]["_v37_aggregates"])
    if b_factors is not None and atoms is not None:
        _ws_b_faktoren(wb, atoms, b_factors, cfg)
    # NEW in v1.0.3: Coordination diagnostic sheet — lists each ligand
    # group and the Gaussian-atom indices assigned to it. Makes it
    # trivial to verify that e.g. HOH oxygens are not leaking into a
    # ligand's atom set (which was the v1.0.2 follow-up bug).
    _ws_coordination(wb, coord_info, atoms, runlog)
    _ws_info(wb, results, cfg, coord_info, dist_ref, logname, cluster_info)
    _save(wb, outfile, runlog)



def _ws_coordination(wb, coord_info, atoms, runlog):
    """Sheet 'Coordination': diagnostic listing of which Gaussian atoms
    each ligand group covers.

    For every ligand group in ``coord_info.group_map`` (e.g. "His 255",
    "Cys 207"), this sheet lists every Gaussian-atom index assigned to
    that group together with the atom name, element and coordinates.
    Makes it trivial to audit:

    - "Does His 255 really have 10 heavy atoms and nothing else?"
    - "Did a crystal water (HOH) with overlapping residue number sneak
      into a ligand group?" (The v1.0.2 follow-up bug.)
    - "Did the index-mismatch fix correctly assign all heavy atoms of
      each ligand?" (The v1.0.2 root-cause bug.)

    Without this sheet, the only place this information was recorded
    was the ``[i] Group '...': N atoms assigned`` lines in REPORT.txt,
    which only gave the count, not the identity. New in v1.0.3.

    Parameters
    ----------
    wb : Workbook
        Target workbook.
    coord_info : CoordInfo
        Coordination info; uses ``.ligands``, ``.group_map`` and the
        ligand residue type / number metadata.
    atoms : list or None
        Gaussian atom list (index -> dict with x, y, z, element, name).
        Indices match Gaussian center IDs in ``group_map``.
    runlog : RunLog
        For warning messages.
    """
    try:
        ws = wb.create_sheet("Coordination")
        C_HEAD = "5D4037"
        ws.row_dimensions[1].height = 24

        # Header line
        hdrs = [
            ("Ligand",   16),
            ("Residue",  12),
            ("Res#",     8),
            ("Coord. atom (donor)", 18),
            ("Fe-X (A)",  10),
            ("# heavy atoms", 14),
            ("Gaussian indices", 60),
        ]
        for ci, (t, w) in enumerate(hdrs, 1):
            _hc(ws, 1, ci, t, w, C_HEAD)

        # Top: per-ligand summary line
        ri = 2
        for lig in (coord_info.ligands or []):
            label   = lig.res_label
            rname   = lig.res_name
            rnum    = lig.res_num
            donor   = f"{lig.lig_element} ({lig.lig_aname})"
            bond_l  = getattr(lig, "bond_len", 0.0)
            centers = coord_info.group_map.get(label, [])
            n_heavy = len(centers)
            idx_str = ", ".join(str(c) for c in sorted(centers))
            _dc(ws, ri, 1, label)
            _dc(ws, ri, 2, rname)
            _dc(ws, ri, 3, rnum)
            _dc(ws, ri, 4, donor)
            _dc(ws, ri, 5, round(float(bond_l), 4) if bond_l else None)
            _dc(ws, ri, 6, n_heavy)
            _dc(ws, ri, 7, idx_str)
            ri += 1

        # Separator
        ri += 1
        ws.cell(ri, 1, "Per-atom detail").font = Font(
            name="Arial", bold=True, size=10, color="5D4037")
        ri += 1

        # Detail header
        det_hdrs = [
            ("Ligand", 16), ("Gaussian #", 12), ("Atom name", 12),
            ("Element", 10), ("x (A)", 10), ("y (A)", 10), ("z (A)", 10),
        ]
        for ci, (t, w) in enumerate(det_hdrs, 1):
            _hc(ws, ri, ci, t, w, C_HEAD)
        ws.row_dimensions[ri].height = 22
        ri += 1

        # Per-atom rows
        n_atoms = len(atoms) if atoms is not None else 0
        for lig in (coord_info.ligands or []):
            label   = lig.res_label
            centers = coord_info.group_map.get(label, [])
            for c in sorted(centers):
                _dc(ws, ri, 1, label)
                _dc(ws, ri, 2, c)
                if atoms is not None and 0 <= c - 1 < n_atoms:
                    a = atoms[c - 1]   # Gaussian centers are 1-indexed
                    # 'atoms' entries usually look like dicts with keys
                    # 'element','x','y','z' (depending on parse). Use safe getters.
                    if isinstance(a, dict):
                        elem  = a.get("element", "?")
                        aname = a.get("name", a.get("aname", "?"))
                        x = a.get("x", 0.0); y = a.get("y", 0.0); z = a.get("z", 0.0)
                    else:
                        # Tuple/list fallback (element, x, y, z) or similar
                        elem  = "?"; aname = "?"
                        x = y = z = 0.0
                    _dc(ws, ri, 3, aname)
                    _dc(ws, ri, 4, elem)
                    _dc(ws, ri, 5, round(float(x), 4))
                    _dc(ws, ri, 6, round(float(y), 4))
                    _dc(ws, ri, 7, round(float(z), 4))
                else:
                    _dc(ws, ri, 3, "(index out of range)")
                ri += 1

        ws.freeze_panes = "A2"

    except Exception as e:
        if runlog is not None:
            runlog.warn(f"Coordination sheet: {e}")


def _ws_b_faktoren(wb, atoms, b_factors, cfg):
    """Sheet: Atomare Debye-Waller-Faktoren from the normal modenanalyse.

    B_i = 8π^2 * (1/3) * Σ_l u_rms(l)^2 * |e_{i,l}|^2  [A^2]

    Nur contributions from the analysierten modes (frequency window
    cfg.freq_min - cfg.freq_max). For Teilfenstern is B_i kleiner
    als the kristallographische B-Faktor.
    """
    ws = wb.create_sheet("B_factors")
    hdrs = ["Atom-Nr.", "element", "x (A)", "y (A)", "z (A)",
            "B_calc (A^2)", "Hinweis"]
    C_B = "004D40"
    for ci, h in enumerate(hdrs, 1):
        _hc(ws, 1, ci, h, 11, C_B)
    ws.row_dimensions[1].height = 24
    ws.column_dimensions["A"].width = 9
    ws.column_dimensions["B"].width = 8
    for col in ["C","D","E","F"]: ws.column_dimensions[col].width = 11
    ws.column_dimensions["G"].width = 44

    freq_note = ""
    if cfg.freq_min is not None or cfg.freq_max is not None:
        lo = f"{cfg.freq_min:.1f}" if cfg.freq_min is not None else "0"
        hi = f"{cfg.freq_max:.1f}" if cfg.freq_max is not None else "∞"
        freq_note = f"Nur modes {lo}-{hi} cm⁻^1"

    for ai, a in enumerate(atoms):
        ri = ai + 2
        b  = float(b_factors[ai]) if ai < len(b_factors) else 0.
        ws.cell(ri, 1, a.get("center",  ai+1)).font = Font(name="Arial", size=9)
        ws.cell(ri, 2, a.get("symbol",  "?")).font  = Font(name="Arial", size=9)
        _dc(ws, ri, 3, round(a["x"], 5))
        _dc(ws, ri, 4, round(a["y"], 5))
        _dc(ws, ri, 5, round(a["z"], 5))
        _dc(ws, ri, 6, round(b,      6))
        if ai == 0 and freq_note:
            ws.cell(ri, 7, freq_note).font = Font(name="Arial", size=8,
                                                   italic=True, color="888888")
    ws.freeze_panes = "A2"


# ===========================================================================
# 2) SSE-EXCEL
# ===========================================================================

def export_sse_excel(
        results:      List[Dict],
        outfile:      str,
        cfg:          Config,
        runlog:       RunLog,
        sse_umap_data: Optional[Tuple] = None,
):
    """Writes the SSE-Analyse in ``_analysis_SSE.xlsx``.

    Wird only erstellt if mindestens a Mode SSE-Daten enthaelt.

    Parameters
    ----------
    results : list of dict
        Modenanalyse-Ergebnisse (required ``results[i]["sse"]``).
    outfile : str
        Zieldatei.
    cfg : Config
        Konfiguration.
    runlog : RunLog
        Fuer Warnmeldungen and filepfad-Registrierung.
    sse_umap_data : tuple, optional
        Ausgabe von ``compute_sse_umap_cluster``; erzeugt
        ``SSE_UMAP_Cluster``-Sheet.

    Notes
    -----
    Sheets: SSE_amplitude_mean | SSE_amplitude_max | SSE_com_amplitude |
    SSE_lateral_std | SSE_lateral_amplitude | SSE_stretching | SSE_axial_amplitude |
    SSE_tilting_angle | SSE_internal_amplitude | SSE_UMAP_Cluster (optional).
    """
    if not [r for r in results if r.get("sse")]:
        runlog.warn(
            "SSE-Export: Kein 'sse'-key in Ergebnissen — "
            f"'{outfile}' will not be created. "
            "Possible causes: analyze_sse=False, PDB without HELIX/SHEET records "
            "und DSSP/phi-psi without Treffer, or alle modes fehlgeschlagen.")
        return   # no SSE-Daten -> no file
    if not _HAS_OPENPYXL:
        runlog.warn("SSE export: openpyxl not installed -- no Excel export possible.")
        return
    wb = Workbook(); wb.remove(wb.active)
    _ws_sse(wb, results, cfg.show_errors_in_excel)
    if sse_umap_data and sse_umap_data[0] is not None:
        _ws_sse_umap(wb, results, sse_umap_data, runlog)
    _save(wb, outfile, runlog)


# ===========================================================================
# 3) EMBEDDINGS-EXCEL
# ===========================================================================

def export_embedding_excel(
        results:          List[Dict],
        embedding_coords: Dict,
        embed_feat_matrix,
        embed_feat_names: List[str],
        cluster_data:     Dict,
        ca_data:          Optional[Tuple],
        outfile:          str,
        cfg:              Config,
        runlog:           RunLog,
        ca_umap_data:     Optional[Tuple] = None,
):
    """Writes Embeddings and Cluster in ``_analysis_Embeddings.xlsx``.

    Parameters
    ----------
    results : list of dict
        Modenanalyse-Ergebnisse.
    embedding_coords : dict of {str: ndarray of shape (n_modes, 2)}
        2D-Koordinaten per Methode.
    embed_feat_matrix : ndarray of shape (n_modes, n_features)
        Feature-Matrix for Cluster-Spalten (Bugfix B4).
    embed_feat_names : list of str
        Feature-Namen.
    cluster_data : dict
        HDBSCAN-Ergebnisse from ``compute_embeddings``.
    ca_data : tuple or None
        ``(ca_centers, ca_res_nums, ca_matrix)`` from Calpha-Analyse.
    outfile : str
        Zieldatei.
    cfg : Config
        Konfiguration.
    runlog : RunLog
        Fuer Warnmeldungen and filepfad-Registrierung.
    """
    if not embedding_coords and not ca_data:
        runlog.warn(
            f"Embeddings export: neither embedding_coords nor ca_data available — "
            f"'{outfile}' will not be created. "
            "compute_embeddings() must be called before export_embedding_excel().")
        return
    wb = Workbook(); wb.remove(wb.active)
    if embedding_coords:
        _ws_projektionen(wb, results, embedding_coords)
    if ca_data:
        _ws_ca(wb, results, ca_data, runlog)
    if cluster_data and embed_feat_names and embed_feat_matrix is not None:
        _ws_cluster(wb, results, cluster_data, embedding_coords,
                    embed_feat_matrix, embed_feat_names)
        _ws_cluster_profil(wb, results, cluster_data, embed_feat_names)
    # NEW in v1.0.3: Ca-UMAP cluster sheets, written into the same
    # _analysis_Embeddings.xlsx alongside the global UMAP and Ca_amplitudes.
    # Placing it here (instead of in _analysis_SSE.xlsx alongside SSE_UMAP)
    # keeps all Ca-related data together: Ca_amplitudes (raw) and
    # Ca_UMAP_clusters / Ca_UMAP_profile (UMAP-derived) in one workbook.
    if ca_umap_data is not None and ca_umap_data[0] is not None:
        _ws_ca_umap(wb, results, ca_umap_data, runlog)
    _save(wb, outfile, runlog)


# ===========================================================================
# 5) INTERPOLATIONS-EXCEL
# ===========================================================================

def export_interpolated_excel(
        results:              List[Dict],
        coord_info:           CoordInfo,
        outfile:              str,
        cfg:                  Config,
        runlog:               RunLog,
        context_results:      Optional[List[Dict]] = None,
        context_results_left: Optional[List[Dict]] = None,
):
    """Writes interpolierte Kennzahlen in ``_analysis_interp{step}.xlsx``.

    Alle Kennzahlen are on ein gleichmaessiges Frequenzraster mit
    Schrittweite ``cfg.interp_step`` interpoliert.
    Enthaelt: OOP/INP, Kern-Groessen, Kern-Scores (10 Symmetrietypen),
    Gruppen-Amplituden, Fe-S-Stretch/Bend, Fe-N-Stretch/Bend.

    Parameters
    ----------
    results : list of dict
        Modenanalyse-Ergebnisse (sortiert after Frequenz).
    coord_info : CoordInfo
        Koordinations-Information.
    outfile : str
        Zieldatei.
    cfg : Config
        Benecessaryt ``interp_step``, ``freq_min``, ``freq_max``,
        ``interp_edge_extend``.
    runlog : RunLog
        Fuer Warnmeldungen and filepfad-Registrierung.
    context_results : list of dict, optional
        Kontext-modes oberhalb von ``freq_max`` (rechts).
    context_results_left : list of dict, optional
        Kontext-modes unterhalb von ``freq_min`` (links).
        Verhindert Nullsetzung if echte modes unterhalb of the Fensters liegen.
    """
    if not _HAS_OPENPYXL:
        runlog.warn("Interpolation export: openpyxl not installed, "
                    "_analysis_interp*.xlsx will not be created.")
        return
    if not results:
        runlog.warn("Interpolation export: no analyzed modes, "
                    "_analysis_interp*.xlsx will not be created.")
        return
    try:
        f_min = cfg.freq_min if cfg.freq_min is not None else \
            min(r["freq"] for r in results) - cfg.interp_edge_extend
        f_max = cfg.freq_max if cfg.freq_max is not None else \
            max(r["freq"] for r in results) + cfg.interp_edge_extend
        f_grid = np.arange(f_min, f_max + cfg.interp_step/2, cfg.interp_step)
        # Kontext-modes an BEIDEN Raendern for korrekte Randinterpolation
        _ctx_l = context_results_left or []
        _ctx_r = context_results or []
        # np.interp requires streng monoton steigende x-values.
        # _all is explizit sortiert -- robust against Raendereffekte und
        # zukuenftige changeen in the Aufrufreihenfolge.
        _all   = sorted(list(_ctx_l) + list(results) + list(_ctx_r),
                        key=lambda r: r["freq"])
        freqs_full = np.array([r["freq"] for r in _all])

        gn   = list(coord_info.group_map.keys())
        ligs_s = sorted(
            {l.res_label for l in coord_info.ligands if l.lig_element == "S"},
            key=lambda s: int(s.split()[-1]) if s.split()[-1].isdigit() else 0)
        ligs_n = sorted(
            {l.res_label for l in coord_info.ligands if l.lig_element == "N"},
            key=lambda s: int(s.split()[-1]) if s.split()[-1].isdigit() else 0)

        bmode = getattr(cfg, "interp_boundary_mode", "context")
        # Warnung if context-Modus without Kontextdaten (Hardening #8)
        # Bugfix v1.0.4 (post-release Apd1 audit): only emit the warning
        # if the user actually requested a frequency window. If no window
        # is set (freq_min and freq_max both None and no freq_windows),
        # the interpolation naturally covers the full DFT spectrum and
        # no context modes are needed. Emitting the warning anyway was
        # misleading.
        _user_set_window = (
            cfg.freq_min is not None
            or cfg.freq_max is not None
            or getattr(cfg, "freq_windows", None) is not None
        )
        if (bmode == "context"
                and not (_ctx_l or _ctx_r)
                and _user_set_window):
            runlog.warn(
                "interp_boundary_mode='context' but no context modes available "
                "-- boundary values set to 0 (wie 'zero'). "
                "interp_context_cm1 increase or 'zero'/'nearest' choose.")

        def ip(vals):
            """Interpoliert on gleichmaessiges Raster.

            Das Randverhalten is through ``cfg.interp_boundary_mode`` gesteuert:

            * ``"context"``  (Standard): Kontext-modes verankern beide Raender.
              Rand-values kommen from echten modes -> physikalisch korrekt.
            * ``"zero"``     : Linker/rechter Rand = 0.0.
              Bedeutung: "no data in this frequency range" (nicht "no motion").
            * ``"nearest"``  : Randwert = naechste presente Mode.
            """
            vals_c = [v if (v is not None and np.isfinite(v)) else 0. for v in vals]
            if bmode == "nearest":
                return np.interp(f_grid, freqs_full, vals_c,
                                 left=vals_c[0] if vals_c else 0.,
                                 right=vals_c[-1] if vals_c else 0.)
            # "context" and "zero": np.interp with left=0/right=0
            # For "context" the Kontext-modes als Anker used
            # (freqs_full enthaelt bereits the Kontext-Frequenzen).
            return np.interp(f_grid, freqs_full, vals_c, left=0.0, right=0.0)

        wb = Workbook(); wb.remove(wb.active)
        ws = wb.create_sheet("Interpolated")

        # Spaltenheader
        his_prot_ligs = [l for l in coord_info.ligands
                         if l.his_protonated and l.res_label in
                         {ll for ll in coord_info.his_ligand_labels}]
        his_prot_labels = sorted({l.res_label for l in his_prot_ligs},
                                  key=lambda s: int(s.split()[-1])
                                  if s.split()[-1].isdigit() else 0)
        headers = (["Frequency (cm-1)",
                    "Lig OOP%", "Lig INP%", "2nd OOP%", "2nd INP%",
                    "Kern OOP%", "Kern |d|(A)", "Kern Lok%", "u_rms(A)"] +
                   [f"Kern: {sk}" for sk in _SCORE_KEYS] +
                   [f"{g} OOP%" for g in gn] + [f"{g} INP%" for g in gn] +
                   [f"{l} Fe-S Stretch" for l in ligs_s] +
                   [f"{l} Fe-S Bend"    for l in ligs_s] +
                   [f"{l} Fe-N Stretch" for l in ligs_n] +
                   [f"{l} Fe-N Bend"    for l in ligs_n] +
                   [f"{l} N-H Stretch"  for l in his_prot_labels])
        for ci, h in enumerate(headers, 1): _hc(ws, 1, ci, h, 12, "1F4E79")
        ws.row_dimensions[1].height = 30

        # Daten interpolieren — globale OOP/INP entfaellt; statt dessen
        # Ring 2 (ligands-Sphaere, lig_oop_pct) als Hauptcharakteristik.
        cols = {
            "lig_oop":    ip([r["lig_oop_pct"]    for r in _all]),
            "lig_inp":    ip([r["lig_inp_pct"]    for r in _all]),
            "second_oop": ip([r["second_oop_pct"] for r in _all]),
            "second_inp": ip([r["second_inp_pct"] for r in _all]),
            "koop":       ip([r["kern_oop"]       for r in _all]),
            "kd":         ip([r["kern_d"]         for r in _all]),
            "kloc":       ip([r.get("kern_loc", 0.) * 100. for r in _all]),
            "urms":       ip([r["u_rms"]          for r in _all]),
        }
        for sk in _SCORE_KEYS:
            cols[f"ks_{sk}"] = ip(
                [r.get("kern_scores", {}).get(sk, 0.) for r in _all])
        for g in gn:
            cols[f"{g}_o"] = ip([r["groups"].get(g,{}).get("oop",0.) for r in _all])
            cols[f"{g}_i"] = ip([r["groups"].get(g,{}).get("inp",0.) for r in _all])
        for l in ligs_s:
            cols[f"{l}_sse"] = ip([r.get("fe_lig",{}).get(l,{}).get("stretch",0.) for r in _all])
            cols[f"{l}_sb"] = ip([r.get("fe_lig",{}).get(l,{}).get("bend",   0.) for r in _all])
        for l in ligs_n:
            cols[f"{l}_ns"] = ip([r.get("fe_lig",{}).get(l,{}).get("stretch",0.) for r in _all])
            cols[f"{l}_nb"] = ip([r.get("fe_lig",{}).get(l,{}).get("bend",   0.) for r in _all])
        for l in his_prot_labels:
            cols[f"{l}_hn"] = ip([r.get("his_hn",{}).get(l,{}).get("hn_stretch",0.)
                                   for r in _all])

        order = (["lig_oop","lig_inp","second_oop","second_inp",
                  "koop","kd","kloc","urms"] +
                 [f"ks_{sk}" for sk in _SCORE_KEYS] +
                 [f"{g}_o" for g in gn] + [f"{g}_i" for g in gn] +
                 [f"{l}_sse" for l in ligs_s] + [f"{l}_sb" for l in ligs_s] +
                 [f"{l}_ns" for l in ligs_n] + [f"{l}_nb" for l in ligs_n] +
                 [f"{l}_hn" for l in his_prot_labels])

        for ri, freq in enumerate(f_grid, 2):
            _dc(ws, ri, 1, float(freq))
            for ci, key in enumerate(order, 2):
                _dc(ws, ri, ci, float(cols[key][ri-2]))

        ws.freeze_panes = "A2"
        n = len(f_grid)
        wb.save(outfile)
        print(f"  -> {os.path.basename(outfile)}"
              f"  ({n} grid points, Δf={cfg.interp_step} cm-1)")
        runlog.add_output(outfile)
    except Exception as e:
        runlog.warn(f"interpolation Excel failed: {e}")



# ===========================================================================
# 5b) SSE-INTERPOLATIONS-EXCEL  (_analysis_SSE_interp.xlsx)
# ===========================================================================

def export_sse_interp_excel(
        results:              List[Dict],
        outfile:              str,
        cfg:                  Config,
        runlog:               RunLog,
        context_results:      Optional[List[Dict]] = None,
        context_results_left: Optional[List[Dict]] = None,
):
    """Writes interpolierte SSE-Amplituden in ``_analysis_SSE_interp{step}.xlsx``.

    Fuer jede the 9 Amplituden-Metriken ein eigenes Sheet.
    Zeilen = SSE-elements, Spalten = interpoliertes Frequenzraster.

    Parameters
    ----------
    results : list of dict
        Modenanalyse-Ergebnisse with ``results[i]["sse"]`` entriesn.
    outfile : str
        Zieldatei.
    cfg : Config
        Benecessaryt ``interp_step``, ``freq_min``, ``freq_max``.
    runlog : RunLog
        Fuer Warnmeldungen and filepfad-Registrierung.
    context_results : list of dict, optional
        Kontext-modes oberhalb von ``freq_max`` (rechts).
    context_results_left : list of dict, optional
        Kontext-modes unterhalb von ``freq_min`` (links).
    """
    if not _HAS_OPENPYXL:
        runlog.warn("SSE interp export: openpyxl not installed, "
                    "_analysis_SSE_interp*.xlsx will not be created.")
        return
    sse_all = [r for r in results if r.get("sse")]
    if not sse_all:
        runlog.warn("SSE-Interp-Export: no modes with SSE data, "
                    "_analysis_SSE_interp*.xlsx will not be created.")
        return

    try:
        freqs  = np.array([r["freq"] for r in sse_all])
        _ctx_l = context_results_left or []
        _ctx_r = context_results or []
        # np.interp requires monoton steigende x-values -> explizit sortieren
        freqs_all_unsorted = (
            [r["freq"] for r in _ctx_l] +
            list(freqs) +
            [r["freq"] for r in _ctx_r]
        )
        sort_idx   = sorted(range(len(freqs_all_unsorted)),
                            key=lambda i: freqs_all_unsorted[i])
        freqs_full = np.array([freqs_all_unsorted[i] for i in sort_idx])
        f_min  = cfg.freq_min if cfg.freq_min is not None else \
            freqs.min() - cfg.interp_edge_extend
        f_max  = cfg.freq_max if cfg.freq_max is not None else \
            freqs.max() + cfg.interp_edge_extend
        f_grid = np.arange(f_min, f_max + cfg.interp_step/2, cfg.interp_step)
        n      = len(f_grid)

        sse_names = list(sse_all[0]["sse"].keys())
        metrics  = [
            "amplitude_mean", "amplitude_max", "com_amplitude",
            "lateral_std",    "lateral_amplitude",  "stretching",
            "axial_amplitude","tilting_angle", "internal_amplitude",
        ]

        wb = Workbook(); wb.remove(wb.active)
        # interp_boundary_mode einmal bestimmen, not in jeder Schleifeniterration
        _sse_bmode = getattr(cfg, "interp_boundary_mode", "context")

        for metric in metrics:
            ws = wb.create_sheet(metric[:31])   # Excel max sheet name length
            _hc(ws, 1, 1, "SSE-element", 20, "1B5E20")
            for ci, freq in enumerate(f_grid, 2):
                ws.cell(1, ci, round(float(freq), 3)).font = \
                    Font(name="Arial", size=8)
                ws.column_dimensions[get_column_letter(ci)].width = 8
            ws.row_dimensions[1].height = 28
            ws.column_dimensions["A"].width = 22

            for ri, sname in enumerate(sse_names, 2):
                ws.cell(ri, 1, sname).font = Font(name="Arial", bold=True, size=9)
                # values for dieses SSE-element over alle Moden
                raw_vals = [r["sse"].get(sname, {}).get(metric, 0.)
                            for r in sse_all]
                raw_c = [v if np.isfinite(v) else 0. for v in raw_vals]
                _left_v  = raw_c[0]  if raw_c else 0.
                _right_v = raw_c[-1] if raw_c else 0.
                if _ctx_l or _ctx_r:
                    raw_full = np.concatenate([
                        np.zeros(len(_ctx_l)), raw_c, np.zeros(len(_ctx_r))])
                    _lv = _left_v  if _sse_bmode == "nearest" else 0.
                    _rv = _right_v if _sse_bmode == "nearest" else 0.
                    interpolated = np.interp(f_grid, freqs_full, raw_full,
                                             left=_lv, right=_rv)
                else:
                    _lv = _left_v  if _sse_bmode == "nearest" else 0.
                    _rv = _right_v if _sse_bmode == "nearest" else 0.
                    interpolated = np.interp(f_grid, freqs, raw_c,
                                             left=_lv, right=_rv)
                for ci, val in enumerate(interpolated, 2):
                    _dc(ws, ri, ci, float(val))

            ws.freeze_panes = "B2"

        wb.save(outfile)
        print(f"  -> {os.path.basename(outfile)}"
              f"  ({n} grid points, {len(sse_names)} SSE-elements, "
              f"{len(metrics)} Metriken)")
        runlog.add_output(outfile)
    except Exception as e:
        runlog.warn(f"SSE-interpolation Excel failed: {e}")




def export_embedding_plots(embedding_coords, results, outname_fn, runlog,
                            sse_umap_data=None, ca_umap_data=None, ca_data=None):
    """Writes Embedding PNGs with frequency, mode-type and cluster coloring.

    Produces (depending on what data is available):

    - ``_embedding_UMAP.png``: global UMAP, 2-panel layout
      (frequency + mode-type).
    - ``_embedding_SSE_UMAP.png``: SSE-feature UMAP, 3-panel layout
      (frequency + mode-type + HDBSCAN cluster). NEW in v1.0.3.
    - ``_embedding_Ca_UMAP.png``: Ca-amplitude UMAP, 3-panel layout.
      NEW in v1.0.3.
    - ``_ca_amplitudes_heatmap.png``: C-alpha amplitudes as a residue ×
      frequency heatmap (log color scale). NEW in v1.0.3.

    Bug fix history: in v1.0.2 and earlier this function only rendered
    ``embedding_coords`` (i.e. the global UMAP), even though SSE-UMAP
    coordinates and C-alpha amplitudes were already computed and
    exported as Excel data. Users running the standard pipeline never
    saw a graphical representation of these results unless they
    plotted them by hand. v1.0.3 adds direct PNG rendering for all
    three secondary embeddings, plus introduces a dedicated Ca-UMAP
    embedding (computed on per-residue C-alpha amplitudes; see
    ``embedding.compute_ca_umap_cluster``). All inputs are optional;
    if absent the corresponding PNG is skipped silently.

    Parameters
    ----------
    embedding_coords : dict of {str: ndarray of shape (n_modes, 2)}
        2D coordinates per method (typically just ``{"UMAP": ...}``).
    results : list of dict
        Mode analysis results.
    outname_fn : callable
        ``cfg.outname``; builds the full output path.
    runlog : RunLog
        For warning messages and output-file registration.
    sse_umap_data : tuple, optional
        Tuple returned by ``embedding.compute_sse_umap_cluster``:
        ``(Z2d, full_labels, feat_names, X_norm, valid_idx, cluster_chars)``.
    ca_umap_data : tuple, optional
        Tuple returned by ``embedding.compute_ca_umap_cluster``:
        ``(Z2d_full, full_labels, feat_names, X_norm, valid_idx, cluster_chars)``.
        ``Z2d_full`` is padded to length ``n_modes`` with NaN rows for
        modes that have no Ca data.
    ca_data : tuple, optional
        Tuple of ``(ca_centers, ca_res_nums, ca_matrix)`` returned by
        ``runner._build_ca_data``. When provided, a heatmap PNG is rendered.
    """
    try:
        import matplotlib; matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except ImportError:
        runlog.warn("matplotlib missing."); return

    freqs=np.array([r["freq"] for r in results])
    _COLS={"Out-of-plane":"#E74C3C","In-plane":"#27AE60","Torsional/Mixed":"#3498DB"}

    def _render_umap_three_panel(Z2d, label_arr, freqs_arr, types_arr,
                                  title_prefix, out_suffix):
        """Helper: renders a 3-panel UMAP figure (freq | mode-type | cluster).

        Z2d shape: (n_pts, 2), label_arr same length, freqs_arr same length,
        types_arr same length. Modes are assumed already filtered to valid.
        """
        fig,(ax1,ax2,ax3)=plt.subplots(1,3,figsize=(20,5))
        sc1=ax1.scatter(Z2d[:,0],Z2d[:,1],c=freqs_arr,cmap="viridis",
                        s=12,alpha=0.7,linewidths=0)
        plt.colorbar(sc1,ax=ax1,label="Frequency (cm-1)")
        ax1.set_title(f"{title_prefix} - frequency")
        ax1.set_xlabel(f"{title_prefix} dim 1"); ax1.set_ylabel(f"{title_prefix} dim 2")
        for mtype,col in _COLS.items():
            mask=[t==mtype for t in types_arr]
            if any(mask):
                ax2.scatter(Z2d[mask,0],Z2d[mask,1],c=col,s=12,alpha=0.7,
                            label=mtype,linewidths=0)
        ax2.set_title(f"{title_prefix} - mode type"); ax2.legend(fontsize=7)
        ax2.set_xlabel(f"{title_prefix} dim 1"); ax2.set_ylabel(f"{title_prefix} dim 2")
        unique_lbls = sorted(set(label_arr))
        try:
            cm = plt.get_cmap("tab10")
        except Exception:
            cm = None
        for li, lbl in enumerate(unique_lbls):
            mask = label_arr == lbl
            n_pts = int(mask.sum())
            if lbl < 0:
                ax3.scatter(Z2d[mask,0],Z2d[mask,1],c="#BBBBBB",s=10,
                            alpha=0.5,linewidths=0,label=f"noise (n={n_pts})")
            else:
                col = cm(li % 10) if cm else None
                ax3.scatter(Z2d[mask,0],Z2d[mask,1],color=col,s=12,
                            alpha=0.8,linewidths=0,
                            label=f"cluster {lbl} (n={n_pts})")
        ax3.set_title(f"{title_prefix} - HDBSCAN cluster")
        ax3.legend(fontsize=6, loc="best")
        ax3.set_xlabel(f"{title_prefix} dim 1"); ax3.set_ylabel(f"{title_prefix} dim 2")
        plt.tight_layout()
        path=outname_fn(out_suffix)
        fig.savefig(path,dpi=150,bbox_inches="tight")
        plt.close(fig)
        print(f"    -> {os.path.basename(path)}")
        runlog.add_output(path)

    # ------------------------------------------------------------------
    # 1) Global UMAP PNGs (one per method in embedding_coords)
    #    Kept at 2-panel layout (no HDBSCAN labels are passed in here).
    # ------------------------------------------------------------------
    for method,Z in embedding_coords.items():
        fig,(ax1,ax2)=plt.subplots(1,2,figsize=(14,5))
        sc1=ax1.scatter(Z[:,0],Z[:,1],c=freqs,cmap="viridis",s=12,alpha=0.7,linewidths=0)
        plt.colorbar(sc1,ax=ax1,label="Frequency (cm-1)")
        ax1.set_title(f"{method} - frequency")
        ax1.set_xlabel(f"{method} dim 1"); ax1.set_ylabel(f"{method} dim 2")
        for mtype,col in _COLS.items():
            mask=[r["mode_type"]==mtype for r in results]
            if any(mask):
                ax2.scatter(Z[mask,0],Z[mask,1],c=col,s=12,alpha=0.7,
                            label=mtype,linewidths=0)
        ax2.set_title(f"{method} - mode type"); ax2.legend(fontsize=7)
        plt.tight_layout()
        path=outname_fn(f"_embedding_{method}.png")
        fig.savefig(path,dpi=150,bbox_inches="tight")
        plt.close(fig)
        print(f"    -> {os.path.basename(path)}")
        runlog.add_output(path)

    # ------------------------------------------------------------------
    # 2) SSE-UMAP 3-panel PNG (NEW in v1.0.3)
    # ------------------------------------------------------------------
    if sse_umap_data is not None:
        try:
            sse_Z2d, sse_full_labels = sse_umap_data[0], sse_umap_data[1]
            sse_valid_idx = sse_umap_data[4] if len(sse_umap_data) > 4 else None
        except (TypeError, IndexError):
            sse_Z2d = sse_full_labels = sse_valid_idx = None

        if sse_Z2d is not None and sse_valid_idx is not None and len(sse_Z2d) > 0:
            freqs_sse = np.array([results[i]["freq"] for i in sse_valid_idx])
            types_sse = [results[i]["mode_type"] for i in sse_valid_idx]
            labels_sse = np.array([sse_full_labels[i] for i in sse_valid_idx])
            _render_umap_three_panel(
                np.asarray(sse_Z2d), labels_sse, freqs_sse, types_sse,
                title_prefix="SSE-UMAP",
                out_suffix="_embedding_SSE_UMAP.png")

    # ------------------------------------------------------------------
    # 3) Ca-UMAP 3-panel PNG (NEW in v1.0.3)
    # ------------------------------------------------------------------
    if ca_umap_data is not None:
        try:
            ca_Z2d_full = ca_umap_data[0]
            ca_full_labels = ca_umap_data[1]
            ca_valid_idx = ca_umap_data[4] if len(ca_umap_data) > 4 else None
        except (TypeError, IndexError):
            ca_Z2d_full = ca_full_labels = ca_valid_idx = None

        if ca_Z2d_full is not None and ca_valid_idx is not None and len(ca_valid_idx) > 0:
            ca_Z2d_full = np.asarray(ca_Z2d_full)
            Z2d_valid = ca_Z2d_full[ca_valid_idx, :]
            freqs_ca = np.array([results[i]["freq"] for i in ca_valid_idx])
            types_ca = [results[i]["mode_type"] for i in ca_valid_idx]
            labels_ca = np.array([ca_full_labels[i] for i in ca_valid_idx])
            _render_umap_three_panel(
                Z2d_valid, labels_ca, freqs_ca, types_ca,
                title_prefix="Ca-UMAP",
                out_suffix="_embedding_Ca_UMAP.png")

    # ------------------------------------------------------------------
    # 4) C-alpha amplitude heatmap PNG (NEW in v1.0.3)
    # ------------------------------------------------------------------
    if ca_data is not None:
        try:
            ca_centers, ca_res_nums, ca_matrix = ca_data[0], ca_data[1], ca_data[2]
        except (TypeError, IndexError):
            ca_centers = ca_res_nums = ca_matrix = None

        if ca_matrix is not None and len(ca_matrix) > 0:
            ca_matrix = np.asarray(ca_matrix, dtype=float)
            n_modes_local = len(freqs)
            n_calpha_local = len(ca_res_nums)
            # _build_ca_data returns shape (n_calpha, n_modes); accept the
            # transposed case as well.
            if ca_matrix.shape == (n_modes_local, n_calpha_local):
                heat = ca_matrix.T
            elif ca_matrix.shape == (n_calpha_local, n_modes_local):
                heat = ca_matrix
            else:
                runlog.warn(
                    f"Ca-heatmap: shape mismatch "
                    f"(ca_matrix={ca_matrix.shape}, "
                    f"n_freqs={n_modes_local}, n_calpha={n_calpha_local}); skipped.")
                return

            with np.errstate(invalid="ignore", divide="ignore"):
                heat_disp = np.log10(np.maximum(heat, 1e-6))

            fig, ax = plt.subplots(1, 1, figsize=(14, 6))
            im = ax.imshow(heat_disp, aspect="auto", origin="lower",
                            cmap="magma",
                            extent=[freqs.min(), freqs.max(), 0, n_calpha_local],
                            interpolation="nearest")
            cbar = plt.colorbar(im, ax=ax)
            cbar.set_label("log10(C-alpha amplitude / A)")
            ax.set_xlabel("Frequency (cm-1)")
            ax.set_ylabel("C-alpha residue index (sequential)")
            ax.set_title("C-alpha amplitudes across frequency spectrum")

            n_ticks = min(15, n_calpha_local)
            tick_idx = np.linspace(0, n_calpha_local-1, n_ticks).astype(int)
            ax.set_yticks(tick_idx)
            ax.set_yticklabels([str(ca_res_nums[i]) for i in tick_idx],
                                fontsize=8)

            plt.tight_layout()
            path=outname_fn("_ca_amplitudes_heatmap.png")
            fig.savefig(path,dpi=150,bbox_inches="tight")
            plt.close(fig)
            print(f"    -> {os.path.basename(path)}")
            runlog.add_output(path)


# ===========================================================================
# Sheet-Implementierungen
# ===========================================================================

def _ws_modenanalyse(wb, results, E, cfg=None):
    """Sheet 'Modenanalyse' (v3.5): kompakte Tabelle with Farbcodierung.

    Ersetzt the fruehere globale OOP/INP through zwei Atom-Ringe:
    Lig-OOP (Cluster-ligands-bond Atome) and 2nd-OOP (Sekundaer-
    Sphaere = volle ligands-Reste + PCET-acceptoren). Die alte
    "Kern OOP"-Spalte (Ring 1 = 4 Cluster-Atome) bleibt erhalten.

    Zellen-Faerbung per value after Signifikanz:
        rot   = trivial      (|X| <= sigma)
        gelb  = signifikant  (sigma < |X| <= 3*sigma)
        gruen = hoch         (|X| > 3*sigma)

    Mode type-Zeilenfaerbung bleibt for "Sym."- and "Type"-Spalten
    erhalten; signifikanz-gefaerbte Zellen ueberschreiben die
    Zeilenfarbung lokal.

    ``E`` is True if error propagation active (Sigma-Spalten are angezeigt).
    """
    thr_low  = float(getattr(cfg, "significance_threshold_low",  1.0)) if cfg else 1.0
    thr_high = float(getattr(cfg, "significance_threshold_high", 3.0)) if cfg else 3.0

    ws=wb.create_sheet("Mode_analysis"); ci=1
    def H(t, w, fg="1F4E79"):
        """Adds a Kopfzelle in the naechsten Spalte ein and erhoet the Zaehler."""
        nonlocal ci; _hc(ws, 1, ci, t, w, fg); ci += 1
    ws.row_dimensions[1].height=36
    H("Mode\n#",8); H("Frequency\n(cm-1)",12); H("u_rms\n(A)",11)
    H("Red.mass\n(AMU)",11); H("Frc.const.\n(mDyn/A)",11)
    H("Sym.",7); H("Type",16); H("Type (fine)",18); H("Prec.",9)
    # Ring 2: ligand bonding atoms
    H("Lig OOP%",10); (H("±sigma",8) if E else None)
    H("Lig INP%",10); (H("±sigma",8) if E else None)
    H("Lig |d|(A)",10); (H("±sigma",8) if E else None)
    # Ring 3: secondary sphere
    H("2nd OOP%",10); (H("±sigma",8) if E else None)
    H("2nd INP%",10); (H("±sigma",8) if E else None)
    H("2nd |d|(A)",10); (H("±sigma",8) if E else None)
    # Ring 1: cluster core
    H("Core OOP%",9); (H("±sigma",8) if E else None)
    H("Core |d|(A)",10); (H("±sigma",8) if E else None)
    H("Core loc%",9)
    H("COM(A)",9); H("Exp.(A)",9); H("Rot.(A)",9)
    # v3.7: three most important lambda columns from reorg_per_mode (Marcus-Hush
    # reorganization energy per mode, mode-mu convention)
    H("Lambda_FeFe\n(cm-1)",13, "2C7A7B")
    H("Lambda_NH\n(cm-1)",13, "2C7A7B")
    H("Lambda_HA\n(cm-1)",13, "2C7A7B")
    n_cols=ci-1
    for ri,r in enumerate(results,2):
        rf=_MTYPE_FILL.get(r["mode_type"]); c=1
        # v3.7: no Klassifikations-Faerbung mehr for Frequenz-Zelle
        # (ohne Score-Schwellen entfaellt the Hierarchie CPET/PT/ET).
        # Mode type-Faerbung is the einzige Hervorhebung.
        freq_fill = None
        reorg_pm = r.get("reorg_per_mode", {}) or {}

        # Linke Spalten (ohne Signifikanz-Faerbung)
        for col_i, v in enumerate([r["number"],r["freq"],r["u_rms"],r["red_mass"],
                  r["frc_const"],r["sym"],r["mode_type"],
                  r.get("mode_type_detail", r["mode_type"]),
                  r["precision"]]):
            _dc(ws,ri,c,v,rf); c+=1
        # Ring 2: Lig OOP / INP / d (mit Signifikanz-Faerbung)
        _dc_sig(ws,ri,c, r.get("lig_oop_pct", 0.), r.get("s_lig_oop", 0.),
                rf, thr_low, thr_high); c+=1
        if E: _sc(ws,ri,c, r.get("s_lig_oop", 0.), rf); c+=1
        _dc_sig(ws,ri,c, r.get("lig_inp_pct", 0.), r.get("s_lig_oop", 0.),
                rf, thr_low, thr_high); c+=1
        if E: _sc(ws,ri,c, r.get("s_lig_oop", 0.), rf); c+=1
        _dc_sig(ws,ri,c, r.get("lig_d", 0.),       r.get("s_lig_d",   0.),
                rf, thr_low, thr_high); c+=1
        if E: _sc(ws,ri,c, r.get("s_lig_d",   0.), rf); c+=1
        # Ring 3: 2nd OOP / INP / d
        _dc_sig(ws,ri,c, r.get("second_oop_pct", 0.), r.get("s_second_oop", 0.),
                rf, thr_low, thr_high); c+=1
        if E: _sc(ws,ri,c, r.get("s_second_oop", 0.), rf); c+=1
        _dc_sig(ws,ri,c, r.get("second_inp_pct", 0.), r.get("s_second_oop", 0.),
                rf, thr_low, thr_high); c+=1
        if E: _sc(ws,ri,c, r.get("s_second_oop", 0.), rf); c+=1
        _dc_sig(ws,ri,c, r.get("second_d", 0.),       r.get("s_second_d",   0.),
                rf, thr_low, thr_high); c+=1
        if E: _sc(ws,ri,c, r.get("s_second_d",   0.), rf); c+=1
        # Ring 1: Kern OOP / d
        _dc_sig(ws,ri,c, r["kern_oop"],  r.get("s_kern_oop", 0.),
                rf, thr_low, thr_high); c+=1
        if E: _sc(ws,ri,c, r["s_kern_oop"], rf); c+=1
        _dc_sig(ws,ri,c, r["kern_d"],    r.get("s_kern_d",   0.),
                rf, thr_low, thr_high); c+=1
        if E: _sc(ws,ri,c, r["s_kern_d"], rf); c+=1
        _dc(ws,ri,c, r.get("kern_loc",0.)*100., rf); c+=1
        for v in (r["cl_com"],r["cl_exp"],r["cl_rot"]):
            _dc(ws,ri,c,v,rf); c+=1
        # v3.7: drei Lambda-values from reorg_per_mode (Mode-mu-Konvention,
        # Marcus-Hush-summierbar). NaN -> leere Zelle (Origin-tauglich).
        lam_fefe = (reorg_pm.get("FeFe", {}) or {}).get("lambda_mode_cm1", float("nan"))
        lam_nh   = (reorg_pm.get("NH",   {}) or {}).get("lambda_mode_cm1", float("nan"))
        lam_ha   = (reorg_pm.get("HA",   {}) or {}).get("lambda_mode_cm1", float("nan"))
        _dc(ws, ri, c, lam_fefe, rf); c += 1
        _dc(ws, ri, c, lam_nh,   rf); c += 1
        _dc(ws, ri, c, lam_ha,   rf); c += 1
    ws.freeze_panes="A2"
    ws.auto_filter.ref=f"A1:{get_column_letter(n_cols)}1"


def _ws_modenanalyse_voll(wb, results, cfg):
    """Sheet 'Modenanalyse_voll' (v3.5): completee values for Publikation.

    Im Gegensatz to kompakten Tabelle:
      - no Farbcodierung
      - jede Sigma-Spalte explizit als eigene Spalte
      - jede Signifikanz-class explizit als String-Spalte
    Geeignet als Supplementary-Information-Tabelle (CSV-similar).
    Wird only erzeugt if cfg.analysis_full = True.
    """
    thr_low  = float(getattr(cfg, "significance_threshold_low",  1.0))
    thr_high = float(getattr(cfg, "significance_threshold_high", 3.0))
    from .core import classify_significance

    ws = wb.create_sheet("Mode_analysis_full"); ci=1
    def H(t, w, fg="1F4E79"):
        nonlocal ci; _hc(ws, 1, ci, t, w, fg); ci += 1
    ws.row_dimensions[1].height = 36
    H("Mode#", 8); H("freq (cm-1)", 12); H("u_rms (A)", 11)
    H("red_mass (AMU)", 12); H("frc_const (mDyn/A)", 14)
    H("sym", 7); H("mode_type", 16); H("mode_type_detail", 18); H("precision", 10)
    # Pro Ring: value, Sigma, Signifikanz
    for ring_pre, base_keys in [
        ("lig",    [("oop_pct", "s_lig_oop"), ("inp_pct", "s_lig_oop"),
                    ("d",       "s_lig_d")]),
        ("second", [("oop_pct", "s_second_oop"), ("inp_pct", "s_second_oop"),
                    ("d",       "s_second_d")]),
    ]:
        for k, sk in base_keys:
            H(f"{ring_pre}_{k}", 12); H(f"sigma_{ring_pre}_{k}", 14)
            H(f"signif_{ring_pre}_{k}", 14)
    # Ring 1 (Kern)
    for k, sk in [("kern_oop", "s_kern_oop"), ("kern_inp", "s_kern_oop"),
                  ("kern_d",   "s_kern_d")]:
        H(k, 12); H(f"sigma_{k}", 14); H(f"signif_{k}", 14)
    H("kern_loc%", 11)
    H("cl_com (A)", 11); H("cl_exp (A)", 11); H("cl_rot (A)", 11)
    n_cols = ci - 1

    for ri, r in enumerate(results, 2):
        c = 1
        for v in [r["number"], r["freq"], r["u_rms"], r["red_mass"],
                  r["frc_const"], r["sym"], r["mode_type"],
                  r.get("mode_type_detail", r["mode_type"]), r["precision"]]:
            _dc(ws, ri, c, v); c += 1
        # Ring 2 (lig)
        for k, sk in [("lig_oop_pct", "s_lig_oop"), ("lig_inp_pct", "s_lig_oop"),
                      ("lig_d",       "s_lig_d")]:
            v = r.get(k, 0.); s = r.get(sk, 0.)
            kl = classify_significance(v, s, thr_low, thr_high)
            _dc(ws, ri, c, v); c += 1
            _sc(ws, ri, c, s); c += 1
            _dc(ws, ri, c, kl); c += 1
        # Ring 3 (2nd)
        for k, sk in [("second_oop_pct", "s_second_oop"),
                      ("second_inp_pct", "s_second_oop"),
                      ("second_d",       "s_second_d")]:
            v = r.get(k, 0.); s = r.get(sk, 0.)
            kl = classify_significance(v, s, thr_low, thr_high)
            _dc(ws, ri, c, v); c += 1
            _sc(ws, ri, c, s); c += 1
            _dc(ws, ri, c, kl); c += 1
        # Ring 1 (Kern)
        for k, sk in [("kern_oop", "s_kern_oop"), ("kern_inp", "s_kern_oop"),
                      ("kern_d",   "s_kern_d")]:
            v = r.get(k, 0.); s = r.get(sk, 0.)
            kl = classify_significance(v, s, thr_low, thr_high)
            _dc(ws, ri, c, v); c += 1
            _sc(ws, ri, c, s); c += 1
            _dc(ws, ri, c, kl); c += 1
        _dc(ws, ri, c, r.get("kern_loc", 0.) * 100.); c += 1
        for v in (r["cl_com"], r["cl_exp"], r["cl_rot"]):
            _dc(ws, ri, c, v); c += 1
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}1"


def _ws_gruppen(wb, results, gn, E):
    """Sheet 'Gruppen_OOP': OOP-/INP-Fractione for each atomare Gruppe (Cys, His, Backbone).

    Diagnostic (v1.0.2): after writing each sheet, count how many group
    rows have all zero values. If any row is 100% zero, emit a UserWarning
    pointing at the most likely cause (PDB-Gaussian atom mapping). The
    silent fallback ``r["groups"].get(g, {}).get(key, 0.)`` used to mask
    this class of bug entirely; see CHANGELOG and the Apd1 bug report.
    """
    import warnings as _w_grp
    for sh,key,fg in [("Gruppen_OOP","oop","C0392B"),("Gruppen_INP","inp","375623"),
                       ("Gruppen_Winkel","angle","784212"),("Gruppen_Tors","torsion","6A0572")]:
        ws=wb.create_sheet(sh)
        _hc(ws,1,1,"Gruppe",14,fg)
        for ji,r in enumerate(results,2): _hc(ws,1,ji,f"{r['freq']:.2f}",9,fg)
        # Track per-group all-zero rows so we can warn the user (instead
        # of silently writing 0 across all modes).
        all_zero_groups: list = []
        for ri,g in enumerate(gn,2):
            ws.cell(ri,1,g).font=Font(name="Arial",bold=True,size=9)
            row_all_zero = True
            for ji,r in enumerate(results,2):
                val = r["groups"].get(g,{}).get(key,0.)
                _dc(ws,ri,ji,val)
                if row_all_zero and val != 0.0:
                    row_all_zero = False
            # Skip the warning for torsion: torsion values are inherently
            # ~1e-3 to 1e-2, and an empty 'tors' list (no atoms with a
            # well-defined tangent) yields a legitimate 0. The OOP/INP/
            # Winkel sheets are the diagnostic ones.
            if row_all_zero and key != "torsion":
                all_zero_groups.append(g)
        if all_zero_groups:
            _w_grp.warn(
                f"Sheet '{sh}': groups {all_zero_groups} are 100% zero "
                f"across all {len(results)} modes. This usually means "
                f"that the PDB-to-Gaussian atom mapping lost these "
                f"residues. Check the run log for related warnings.",
                UserWarning, stacklevel=2)


def _ws_fe_bindung(wb, results, coord_info, element_filter, E, runlog=None,
                    cfg=None):
    """Fe-X bondanalyse for einen elementtyp.

    Jedes Residue bekommt einen eigenen Abschnittsheader (Residuenname
    einmalig), darunter Stretch (A), Bend (A), Bend INP (A), Bend OOP (A)
    and Bend Signif. als Unterzeilen.

    v3.5: Bend is additionally in INP- and OOP-Fractione bezueglich der
    cluster plane zerplaces. Die Fractione are in Aengstroem with Sigma
    angiven; Zellen are after Signifikanz farbcodiert.

    Parameters
    ----------
    element_filter : str
        ``"S"`` for Cys (Fe-S) or ``"N"`` for His (Fe-N).
    """
    thr_low  = float(getattr(cfg, "significance_threshold_low",  1.0)) if cfg else 1.0
    thr_high = float(getattr(cfg, "significance_threshold_high", 3.0)) if cfg else 3.0

    ligs = [l for l in coord_info.ligands if l.lig_element == element_filter]
    if not ligs:
        if runlog is not None:
            runlog.info(
                f"_ws_fe_bindung: Keine ligands with element='{element_filter}' — "
                f"Sheet will not be created (normal für reine Cys4- or His-freie Cluster).")
        return

    # Residuen sortiert after Nummer
    labels = sorted(
        {l.res_label for l in ligs},
        key=lambda s: int(s.split()[-1]) if s.split()[-1].isdigit() else 0)

    if element_filter == "S":
        sheet_name, color, bond_title = "Fe_S_Cys", "1A5276", "Fe-S-bonden (Cys)"
    else:
        sheet_name, color, bond_title = "Fe_N_His", "6A0572", "Fe-N-bonden (His)"

    ws = wb.create_sheet(sheet_name)
    ws.row_dimensions[1].height = 28
    _hc(ws, 1, 1, bond_title, 24, color)
    for ji, r in enumerate(results, 2):
        _hc(ws, 1, ji, f"{r['freq']:.2f}", 9, color)

    row = 2
    for lbl in labels:
        # Abschnittsheader: Residuenname EINMALIG in dunkler Farbe
        c = ws.cell(row, 1, lbl)
        c.font      = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor=color)
        c.alignment = Alignment(horizontal="left", vertical="center")
        b = _bd()
        if b: c.border = b
        row += 1

        # values with Signifikanz-Faerbung (rot/gelb/gruen)
        # Stretch (A) -- Sigma-key is "s_stretch"
        # Bend total (A) -- "s_bend"
        # Bend INP (A) -- "s_bend_inp"
        # Bend OOP (A) -- "s_bend_oop"
        # Plus Signifikanz-classes-Zeile (String)
        rows_spec = [
            ("  Stretch (A)",  "stretch",  "s_stretch"),
            ("  Bend (A)",     "bend",     "s_bend"),
            ("  Bend INP (A)", "bend_inp", "s_bend_inp"),
            ("  Bend OOP (A)", "bend_oop", "s_bend_oop"),
        ]
        for display, vk, sk in rows_spec:
            ws.cell(row, 1, display).font = Font(name="Arial", size=9)
            for ji, r in enumerate(results, 2):
                lig = r.get("fe_lig", {}).get(lbl, {})
                v   = lig.get(vk, 0.)
                s   = lig.get(sk, 0.)
                _dc_sig(ws, row, ji, v, s, None, thr_low, thr_high)
            row += 1
            if E:
                ws.cell(row, 1, "    ±sigma").font = Font(
                    name="Arial", size=8, italic=True, color="888888")
                for ji, r in enumerate(results, 2):
                    _sc(ws, row, ji,
                        r.get("fe_lig", {}).get(lbl, {}).get(sk, 0.))
                row += 1

        # Signifikanz-class for bend_oop vs bend_inp (Differenz-Signifikanz)
        ws.cell(row, 1, "  Bend Signif.").font = Font(
            name="Arial", size=8, italic=True, color="555555")
        for ji, r in enumerate(results, 2):
            lig = r.get("fe_lig", {}).get(lbl, {})
            sig = lig.get("bend_significance", "")
            cell = ws.cell(row, ji)
            cell.value = sig
            cell.font = Font(name="Arial", size=8, italic=True, color="555555")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            # Farbcodierung the Klassifikations-Zelle
            if sig and sig != "trivial":
                if "high" in sig:    cell.fill = PatternFill("solid", fgColor="E0F0E0")
                elif "significant" in sig: cell.fill = PatternFill("solid", fgColor="FFF8E0")
            elif sig == "trivial":
                cell.fill = PatternFill("solid", fgColor="FFE0E0")
        row += 1

    ws.freeze_panes = "B2"

    # v1.0.4: all-zero-row detector, identical pattern to _ws_gruppen.
    # If for some ligand all of stretch/bend/bend_inp/bend_oop are 0
    # across every mode, the most likely cause is a c2l lookup failure
    # in analyze_fe_ligand that triggered the silent _zero_lig() fallback
    # (the v1.0.4 audit revealed this bug class). Emit a UserWarning so
    # the user notices the dead row immediately. The UserWarning from
    # core.analyze_fe_ligand will fire first; this is the downstream
    # double-check at the export layer.
    import warnings as _w_fe
    all_zero_ligs: list = []
    for lbl in labels:
        is_zero = True
        for r in results:
            lig = r.get("fe_lig", {}).get(lbl, {})
            for vk in ("stretch", "bend", "bend_inp", "bend_oop"):
                if lig.get(vk, 0.) != 0.0:
                    is_zero = False
                    break
            if not is_zero:
                break
        if is_zero:
            all_zero_ligs.append(lbl)
    if all_zero_ligs:
        _w_fe.warn(
            f"Sheet '{sheet_name}': ligands {all_zero_ligs} are 100% "
            f"zero across all {len(results)} modes (stretch/bend "
            f"identically zero). This usually means analyze_fe_ligand "
            f"hit a c2l lookup failure and returned _zero_lig() "
            f"silently. Check for related warnings from "
            f"analyze_fe_ligand earlier in the run.",
            UserWarning, stacklevel=2)

    # v3.5: Vollstaendiges Sheet (additionally) for analysis_full = True
    if cfg is not None and getattr(cfg, "analysis_full", False):
        _ws_fe_bindung_voll(wb, results, coord_info, element_filter, cfg)


def _ws_fe_bindung_voll(wb, results, coord_info, element_filter, cfg):
    """Vollstaendiges Sheet 'Fe_S_Cys_voll' bzw. 'Fe_N_His_voll' (v3.5).

    A Zeile per (Mode x ligands)-Kombination, alle values and Sigmas
    explizit als Spalten. Ideal for SI-Tabellen and Paper-Reproduktion.
    """
    from .core import classify_significance
    thr_low  = float(getattr(cfg, "significance_threshold_low",  1.0))
    thr_high = float(getattr(cfg, "significance_threshold_high", 3.0))

    ligs = [l for l in coord_info.ligands if l.lig_element == element_filter]
    if not ligs:
        return
    labels = sorted({l.res_label for l in ligs},
                    key=lambda s: int(s.split()[-1]) if s.split()[-1].isdigit() else 0)
    sheet_name = "Fe_S_Cys_voll" if element_filter == "S" else "Fe_N_His_voll"
    color      = "1A5276"        if element_filter == "S" else "6A0572"

    ws = wb.create_sheet(sheet_name); ci = 1
    def H(t, w, fg=color):
        nonlocal ci; _hc(ws, 1, ci, t, w, fg); ci += 1
    ws.row_dimensions[1].height = 28
    H("Mode#", 8); H("freq (cm-1)", 12); H("ligands", 14); H("element", 9)
    for k in ["stretch", "bend", "bend_inp", "bend_oop"]:
        H(f"{k} (A)", 12); H(f"sigma_{k}", 12); H(f"signif_{k}", 12)
    H("bend_inp_pct", 12); H("bend_oop_pct", 12)
    H("bend_significance", 18)
    n_cols = ci - 1

    ri = 2
    for r in results:
        for lbl in labels:
            lig = r.get("fe_lig", {}).get(lbl, {})
            if not lig:
                continue
            c = 1
            _dc(ws, ri, c, r["number"]); c += 1
            _dc(ws, ri, c, r["freq"]); c += 1
            _dc(ws, ri, c, lbl); c += 1
            _dc(ws, ri, c, lig.get("element", "?")); c += 1
            for k in ["stretch", "bend", "bend_inp", "bend_oop"]:
                v = lig.get(k, 0.)
                s = lig.get(f"s_{k}", 0.)
                kl = classify_significance(v, s, thr_low, thr_high)
                _dc(ws, ri, c, v); c += 1
                _sc(ws, ri, c, s); c += 1
                _dc(ws, ri, c, kl); c += 1
            _dc(ws, ri, c, lig.get("bend_inp_pct", float("nan"))); c += 1
            _dc(ws, ri, c, lig.get("bend_oop_pct", float("nan"))); c += 1
            _dc(ws, ri, c, lig.get("bend_significance", "")); c += 1
            ri += 1
    ws.freeze_panes = "A2"
    if ri > 2:
        ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}1"


def _ws_his_hn(wb, results, his_ligs, E):
    """Sheet 'Fe_N_His': H-N-bondaenderungen for protonierte His-ligands."""
    ws=wb.create_sheet("His_HN")
    _hc(ws,1,1,"His / H-N",22,"6A0572")
    for ji,r in enumerate(results,2): _hc(ws,1,ji,f"{r['freq']:.2f}",9,"6A0572")
    row=2; done=set()
    for lig in his_ligs:
        lb=lig.res_label
        if lb in done: continue
        done.add(lb)
        ws.cell(row,1,f"{lb} H-N Stretch(A)").font=Font(name="Arial",bold=True,size=9)
        for ji,r in enumerate(results,2):
            _dc(ws,row,ji,r.get("his_hn",{}).get(lb,{}).get("hn_stretch",0.))
        row+=1
        if E:
            ws.cell(row,1,"  ±sigma").font=Font(name="Arial",size=8,italic=True,color="888888")
            for ji,r in enumerate(results,2):
                _sc(ws,row,ji,r.get("his_hn",{}).get(lb,{}).get("s_hn_stretch",0.))
            row+=1
    ws.freeze_panes="B2"

    # v1.0.4: all-zero-row detector for His_HN. Identical pattern to
    # _ws_fe_bindung. For each His ligand we check whether hn_stretch is
    # zero across ALL modes. Caveat: a deprotonated His will yield an
    # empty his_hn dict entry (the get() yields 0.), which is LEGITIMATE
    # silence. We therefore only warn when the ligand IS protonated:
    # for that we look at lig.his_protonated.
    import warnings as _w_hn
    all_zero_hn: list = []
    for lig in his_ligs:
        if not getattr(lig, "his_protonated", False):
            continue  # deprot is legitimate silence -- skip warning
        lb = lig.res_label
        is_zero = True
        for r in results:
            v = r.get("his_hn", {}).get(lb, {}).get("hn_stretch", 0.)
            if v != 0.0:
                is_zero = False
                break
        if is_zero:
            all_zero_hn.append(lb)
    if all_zero_hn:
        _w_hn.warn(
            f"Sheet 'His_HN': protonated His ligands {all_zero_hn} "
            f"have hn_stretch = 0 across all {len(results)} modes. "
            f"This usually means analyze_his_hn skipped the ligand "
            f"silently (c2l lookup failure for N or H center). "
            f"Check for related warnings from analyze_his_hn earlier "
            f"in the run.",
            UserWarning, stacklevel=2)


def _ws_kern_scores(wb, results, E):
    """Sheet 'Kern_Scores': heuristische motionsmuster-Scores aller modes.

    Notes
    -----
    **Heuristik, not orthogonal**: Die Scores are geometrische
    Projektionen the Eigenvectors on qualitative motionsmuster
    (Breathing, Hinge-Fe, ...). Sie are *nicht* the orthogonale
    D2h symmetry decomposition — mehrere Scores koennen demselben Irrep
    correspond to (z. B. are Breathing/Fe-stretching/S-stretching alle
    Ag-Fractione), and einige Scores mischen mehrere Irreps.

    Fuer rigorose, orthogonale Symmetriezerlegung after Kingsbury & Senge
    see the **SCSD-Sheet**. Beide Sheets are komplementaer: die
    Heuristik gibt intuitive motionsmuster (z. B. "Breathing"), das
    SCSD the exakte D2h-Irrep-Aufteilung (Ag, B1g, ...).

    D2h-Irrep-Mapping (vereinfacht):
      * Translation, Umbrella, Hinge-Fe, Hinge-S — sollten ~ 0 sein
        for saubere Eckart-orthogonale Vibrationen
      * Fe-stretching, S-stretching, Breathing — alle Ag (nicht orthogonal)
      * Rotation-ip, Rhombus-Scherung — naeherungsweise B1g
      * OOP-Twist — Mischung from B2g + B3g
    """
    ws = wb.create_sheet("Core_scores")
    # Heuristik-Hinweis in Zeile 1
    from openpyxl.styles import Alignment as _Align
    _note = ("HINWEIS: Kern-Scores are HEURISTISCH (geometrische Projektionen, "
             "not orthogonal, generally do not sum to 1). "
             "Fuer the rigorose D2h symmetry decomposition after Kingsbury & Senge "
             "(Chem. Sci. 15, 13638, 2024) the SCSD-Sheet verwenden.")
    c_note = ws.cell(1, 1, _note)
    c_note.font = Font(name="Arial", italic=True, size=8, color="7B2D8B")
    c_note.alignment = _Align(wrap_text=True)
    n_cols = 4 + len(_SCORE_KEYS)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.row_dimensions[1].height = 26
    hdrs = [("Frequency", 12), ("Type", 14), ("Core mode 1", 14), ("Core mode 2", 14)]
    for sk in _SCORE_KEYS:
        hdrs.append((sk, 12))
    ws.row_dimensions[2].height = 34
    for ci, (t, w) in enumerate(hdrs, 1):
        _hc(ws, 2, ci, t, w)
    for ri, r in enumerate(results, 3):
        rf = _MTYPE_FILL.get(r["mode_type"]); ks = r.get("kern_scores", {})
        _dc(ws, ri, 1, r["freq"], rf); _dc(ws, ri, 2, r["mode_type"], rf)
        _dc(ws, ri, 3, r.get("kern_primary", "n/a"), rf)
        _dc(ws, ri, 4, r.get("kern_secondary", "-"), rf)
        for ci, sk in enumerate(_SCORE_KEYS, 5):
            _dc(ws, ri, ci, float(ks.get(sk, 0.)), rf)


def _ws_abstaende(wb, dist_ref):
    """Sheet 'Info_Abstaende': Referenz-Clustergeometrie (Fe-Fe, Fe-S, S-S in Angstrom)."""
    ws=wb.create_sheet("Equilibrium_distances")
    for ci,(t,w) in enumerate([("Distance",14),("d (A)",14),("±sigma (A)",14)],1):
        _hc(ws,1,ci,t,w)
    for ri,(k,(v,sv)) in enumerate(dist_ref.items(),2):
        ws.cell(ri,1,k).font=Font(name="Arial",bold=True,size=9)
        _dc(ws,ri,2,v); _sc(ws,ri,3,sv)


def _ws_scsd(wb, results, scsd_res):
    """Sheet 'SCSD': rigorose D2h symmetry decomposition the clustergeometrie.

    Methode: Kingsbury & Senge, *Chem. Sci.* **15**, 13638 (2024).
    Kanonische D2h-Referenz (Fe-Fe = 2.73 A, Fe-S = 2.20 A); Achsen
    x = Fe-Fe, y = S-S, z = cluster normal.

    Die Spalten ``Kern-Modus (SCSD)`` zeigen the *rigorose* dominante
    Irrep-Klassifikation (aus ``|SCSD_d<Irr>|``). Das is die
    physikalisch belastbare Antwort on the Frage 'in welche
    Symmetriekoordinate zerfaellt diese Mode?'. Die heuristischen
    motions-Labels (Breathing, Hinge-Fe, ...) finden im
    Kern_Scores-Sheet.
    """
    ws = wb.create_sheet("SCSD")
    # Kopfzeile (Zeile 1): methodshinweis
    from openpyxl.styles import Alignment as _Align
    _note = ("METHOD: Kingsbury & Senge, Chem. Sci. 15, 13638 (2024). "
             "Canonical D2h reference (Fe-Fe=2.73A, Fe-S=2.20A). "
             "Axes: x=Fe-Fe, y=S-S, z=normal. "
             "SCSD values are ORTHOGONAL and directly comparable between structures.")
    c_note = ws.cell(1, 1, _note)
    c_note.font = Font(name="Arial", italic=True, size=8, color="2C5F8E")
    c_note.alignment = _Align(wrap_text=True)

    first = scsd_res[0]["scsd"]
    irreps = first.get("scsd_irreps", [])
    geo_keys = sorted([k[:-4] for k in first if k.endswith("_ref")
                       and not k.startswith("SCSD") and not k.startswith("s_")
                       and not k.startswith("disp") and k not in ("com_disp_cluster","total_ref")])
    # Header: Frequenz, Typ, SCSD-Primary, SCSD-Secondary, Heuristik-Primary, ...
    hdrs = [("Frequency", 12),
            ("Type", 14),
            ("Core mode (SCSD)", 14),
            ("Secondary (SCSD)", 14),
            ("Core mode (heuristic)", 18)]
    for irr in irreps:
        hdrs += [(f"SCSD_{irr}_ref", 10), ("\u00b1sigma", 7),
                 (f"SCSD_{irr}_dist", 10), ("\u00b1sigma", 7),
                 (f"dSCSD_{irr}", 10), ("\u00b1sigma", 7)]
    hdrs += [("Total_ref", 10), ("Total_dist", 10), ("Total_d", 10)]
    for gb in geo_keys:
        hdrs += [(f"{gb}_ref", 10), (f"{gb}_dist", 10),
                 (f"d_{gb}", 10), ("\u00b1sigma", 7)]
    for lb in ["Fe1", "Fe2", "S1", "S2"]:
        hdrs.append((f"disp_{lb}", 10))
    hdrs.append(("COM_disp", 10))

    # Note-Zeile spannt over alle Spalten
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(hdrs))
    ws.row_dimensions[1].height = 26

    # Header in Zeile 2
    ws.row_dimensions[2].height = 34
    for ci, (t, w) in enumerate(hdrs, 1):
        _hc(ws, 2, ci, t, w)

    # data rows ab Zeile 3
    for ri, r in enumerate(results, 3):
        rf = _MTYPE_FILL.get(r["mode_type"]); sc = r.get("scsd", {})
        _dc(ws, ri, 1, r["freq"], rf)
        _dc(ws, ri, 2, r["mode_type"], rf)
        # SCSD-rigorose Klassifikation (neu)
        _dc(ws, ri, 3, sc.get("scsd_primary",   "n/a"), rf)
        _dc(ws, ri, 4, sc.get("scsd_secondary", "-"),   rf)
        # Heuristik-Klassifikation (to Vergleichsanzeige)
        _dc(ws, ri, 5, r.get("kern_primary", "n/a"), rf)
        ci = 6
        for irr in irreps:
            for fk, sfn in [(f"SCSD_{irr}_ref", f"s_SCSD_{irr}_ref"),
                            (f"SCSD_{irr}_dist", f"s_SCSD_{irr}_dist"),
                            (f"SCSD_d{irr}", f"s_SCSD_d{irr}")]:
                _dc(ws, ri, ci, sc.get(fk, 0.), rf); ci += 1
                _sc(ws, ri, ci, sc.get(sfn, 0.), rf); ci += 1
        for fk in ("total_ref", "total_dist", "total_d"):
            _dc(ws, ri, ci, sc.get(f"SCSD_{fk}", sc.get(fk, 0.)), rf); ci += 1
        for gb in geo_keys:
            _dc(ws, ri, ci, sc.get(f"{gb}_ref", 0.),  rf); ci += 1
            _dc(ws, ri, ci, sc.get(f"{gb}_dist", 0.), rf); ci += 1
            _dc(ws, ri, ci, sc.get(f"d_{gb}", 0.),    rf); ci += 1
            _sc(ws, ri, ci, sc.get(f"s_d_{gb}", 0.),  rf); ci += 1
        for lb in ["Fe1", "Fe2", "S1", "S2"]:
            _dc(ws, ri, ci, sc.get(f"disp_{lb}", 0.), rf); ci += 1
        _dc(ws, ri, ci, sc.get("com_disp_cluster", 0.), rf)


def _ws_pcet_et_kandidaten(wb, results, coord_info):
    r"""[DEPRECATED] Altes v3.5-Sheet, in v3.6 ersetzt through _ws_cpet_kandidaten,
    _ws_pt_kandidaten, _ws_et_kandidaten and _ws_reorganisationsenergie.

    Diese Funktion bleibt nur, if aelterer Code sie still calls;
    in v3.6 is sie not mehr used. Die neue Aufruf-Site benutzt
    the drei separaten Sheets plus the Origin-taugliche Tabelle.
    """
    return  # No-op


# ===========================================================================
# v3.6: Drei Score-Sheets + reorganization energy-Sicht (Origin-tauglich)
# ===========================================================================
#
# Architektur:
#   * reorganization energy-Sheet (Origin-Sicht): alle Modes, frequenz-
#     sortiert, a Spalte per Lambda-Komponente, leere Zellen statt NaN-
#     Strings. Direkt for Origin-Plot importierbar.
#   * Drei Kandidaten-Sheets (Anwender-Sicht): only klassifizierte Modes,
#     sortiert after Lambda absteigend, with Farbcodierung and Rahmen.
#   * CPET-pathe-Sheet: Reaktionspfad-Vektoren the Top-CPET-Modes mit
#     vorzeichenbehafteter dr_NH and dr_HA — zeigt the motionsmuster.
#
# Farbpalette (konsistent between allen v3.6-Sheets):
#   CPET     -> Lila    (kraftig: D6BCFA, hell: EDE0FE)
#   PT-only  -> Tuerkis (kraftig: 9DECF9, hell: D6F4F9)
#   ET       -> Orange  (kraftig: F6AD55, hell: FEEBC8)
#   Top-3 Reorg-contribution per class: dicker schwarzer Rahmen (additiv).
# ---------------------------------------------------------------------------

# Farbpalette
_V36_FILL_CPET_STARK     = "D6BCFA"
_V36_FILL_CPET_MODERAT   = "EDE0FE"
_V36_FILL_PT_STARK       = "9DECF9"
_V36_FILL_PT_MODERAT     = "D6F4F9"
_V36_FILL_ET_STARK       = "F6AD55"
_V36_FILL_ET_MODERAT     = "FEEBC8"




def _ws_info(wb, results, cfg, coord_info, dist_ref, logname, cluster_info):
    """Sheet 'Info': Analyseparameter, Clustergeometrie and Laufzusammenfassung."""
    ws=wb.create_sheet("Info"); fl=[r["freq"] for r in results]
    hp=sum(1 for r in results if r["precision"]=="high")

    def ih(row, text, color="1F4E79"):
        """Writes a Info-Ueberschriftszeile (farbig, fett, over 3 Spalten)."""
        c=ws.cell(row,1,text)
        c.font=Font(name="Arial",bold=True,size=9,color="FFFFFF")
        c.fill=PatternFill("solid",fgColor=color)
        ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=3)
        return row+1
    def ir(row, key, val1, val2=""):
        """Writes a dreispaltige Info-Zeile: key | value | Zusatz."""
        ws.cell(row,1,key).font=Font(name="Arial",bold=True,size=9)
        ws.cell(row,2,str(val1)).font=Font(name="Arial",size=9)
        ws.cell(row,3,str(val2)).font=Font(name="Arial",size=9,color="555555")
        return row+1

    ri=1
    ri=ih(ri,"Analyseparameter")
    ri=ir(ri,"Quelldatei",logname)
    ri=ir(ri,"Moden",len(results))
    ri=ir(ri,"frequency range",f"{min(fl):.2f}-{max(fl):.2f} cm-1" if fl else "-")
    ri=ir(ri,"Filter",
          f"{'-' if cfg.freq_min is None else cfg.freq_min} - "
          f"{'-' if cfg.freq_max is None else cfg.freq_max} cm-1")
    ri=ir(ri,"HP-Moden",f"{hp}/{len(results)}")
    ri=ir(ri,"Temperature",f"{cfg.temp_k} K" if cfg.temp_k is not None else "Amplitude mode")
    ri=ir(ri,"PDB-Kette",cfg.pdb_chain or "(alle)")
    ri=ir(ri,"Erstellt",time.strftime("%Y-%m-%d %H:%M"))
    ri+=1
    ri=ih(ri,"[2Fe-2S]-Cluster Geometrie","375623")
    if cluster_info:
        for lb,ctr,elem,coords,note in cluster_info:
            cs=(f"({coords[0]:+.3f},{coords[1]:+.3f},{coords[2]:+.3f})A" if coords else "-")
            ri=ir(ri,lb,f"Center {ctr} ({elem})",f"{cs} {note}")
        if dist_ref:
            ri+=1
            for dk,(dv,dsv) in dist_ref.items():
                ri=ir(ri,f"  {dk}",f"{dv:.6f} A",f"±{dsv:.1e} A")
    ri+=1
    ri=ih(ri,"Koordinierende amino acidn (automatically erkannt)","7B3F00")
    if coord_info.ligands:
        for lig in sorted(coord_info.ligands,key=lambda l:(l.fe_idx,l.res_num)):
            prot=" [prot., H-N verfuegbar]" if lig.his_protonated else ""
            n=len(coord_info.group_map.get(lig.res_label,[]))
            ri=ir(ri,f"{lig.res_label}{prot}",
                  f"Fe{lig.fe_idx+1}, {lig.lig_element}({lig.lig_aname}), d={lig.bond_len:.3f}A",
                  f"{n} Atome")
    else:
        ri=ir(ri,"(keine erkannt - PDB-Kette and Kabsch-RMSD pruefen)","")
    ri+=1
    ri=ih(ri,"Methodik & Kennzeichnung heuristischer Kenngroessen","7B2D8B")
    ri=ir(ri,"Kern-Scores (Kern_Scores-Sheet)",
          "HEURISTISCH - geometrische motionsmuster (Breathing, Hinge, ...)",
          "Nicht orthogonal; mehrere Scores koennen demselben D2h-Irrep "
          "correspond to (e.g. Breathing/Fe-stretching/S-stretching = alle Ag)")
    ri=ir(ri,"SCSD (SCSD-Sheet)",
          "RIGOROUS - orthogonale D2h symmetry decomposition (Kingsbury-Methode)",
          "values over Strukturen direkt vergleichbar (kanonische Referenz)")
    ri=ir(ri,"SSE-Amplituden (SSE-Sheets)",
          "HEURISTISCH - geometrische Naeherungsmetriken",
          "s_*-Spalten: heuristische Stabilitaetsmasse, no exakte error propagation")
    ri=ir(ri,"OOP/INP-Klassifikation (binaer, Spalte 'Typ')",
          "Out-of-plane: OOP% > 60; In-plane: INP% > 60; sonst Torsional/Mixed",
          "Faerbung the Excel-Zeilen, Plot-Legenden")
    ri=ir(ri,"OOP/INP-Klassifikation (fein, Spalte 'Typ (fein)')",
          "7 Stufen with Schwellen (60, 75, 90): Pur/Stark/Mehrheitlich OOP|INP, Gemischt",
          "PCET-Vergleich: feinere Trennung schwach/stark gemischter Moden")
    ri=ir(ri,"HP/Std-Frequenz-Konsistenzcheck (REPORT)",
          "Vergleich the Frequenzen between HP- and Standard-blocksn (tolerance 0.01 cm-1)",
          "Abweichungen: Parser-Versatz, korrupte file, or verkettete freq-Jobs")
    ri=ir(ri,"Thermodynamische Parameter (REPORT: 'NIS / Fe-pDOS')",
          "ELEMENT-SELEKTIV (Fe-pDOS, per Fe atom): f_LM, <x^2>_Fe, <U>_Fe, <T>_Fe, c_V^Fe, S_Fe, <k>_Fe, <E>_Fe",
          "NICHT the values for clusters, ligands or Protein als Ganzes - "
          "NRVS misst only 57Fe-Kernuebergaenge")
    ri=ir(ri,"Multi-Cluster-Auswahl (cluster_index)",
          "For multi-[2Fe-2S] systems (e.g. glutaredoxin dimer), cluster #cfg.cluster_index is analyzed",
          "default 0 = engstes Fe-Fe-Paar; alle Cluster in the REPORT gelistet")
    ri=ir(ri,"bond reorg energyn (Sheets reorganization energy, Reorg_Total, Reorg_pro_bond)",
          "Pro Mode dr_X and lambda_X = (1/2) mu omega^2 (dr_X)^2 for "
          "X = FeFe, FeN, FeS, NH, HA",
          "HA in the reaction coordinaten-Modus (v3.7.1); thermisch scalede "
          "Mode-Amplituden (see Handbuch §sec:reorg)")
    ri=ir(ri,"PCET-Reorg (Lambda_NH, Lambda_HA)",
          "Aktiv NUR with His-ligands + H-acceptoren in Cutoff-Reichweite",
          "For rein Cys-koordinierten Systemen or deprot. His are "
          "Lambda_NH and Lambda_HA = 0 (Kanal entfaellt)")
    ri=ir(ri,"Modulations- and kumulative Spektren (Sheets Modulations_Spektren, Lambda_kumulativ)",
          "M_X(omega) = sum_i |dr_X(i)| * Gauss(omega - omega_i, sigma); "
          "Lambda_X(omega_cut) = sum_{omega_i <= omega_cut} lambda_X(i)",
          "Direkter Vergleich with NRVS-Banden in Origin")
    ri=ir(ri,"SCSD-Methode",
          "Kingsbury & Senge, Chem. Sci. 15, 13638 (2024)",
          "https://www.kingsbury.id.au/scsd  -  Python: scsdpy")
    ri=ir(ri,"SCSD-Referenzgeometrie",
          "KANONISCH FEST: Fe-Fe=2.73 A, Fe-S=2.20 A, D2h-Rhombus",
          "Meane Rieske/Ferredoxin-Kristallstrukturen; "
          "Achsen: x=Fe-Fe, y=S-S, z=cluster normal")
    ri=ir(ri,"Anwendungsbereich",
          "Optimiert for [2Fe-2S]-Rieske/Ferredoxin-Systeme",
          "Andere Metallcluster: Cutoffs and Referenzgeometrie anpassen")
    ws.column_dimensions["A"].width=35
    ws.column_dimensions["B"].width=45
    ws.column_dimensions["C"].width=55


def _ws_sse(wb, results, E):
    """Sheet 'SSE_elements': Amplituden aller Sekundaerstrukturelemente per Mode."""
    sse_all=[r for r in results if r.get("sse")]
    if not sse_all: return
    sse_names=list(sse_all[0]["sse"].keys())
    for metric in ["amplitude_mean","amplitude_max","com_amplitude","lateral_std",
                   "lateral_amplitude","stretching","axial_amplitude","tilting_angle",
                   "internal_amplitude"]:
        ws=wb.create_sheet("SSE_"+metric[:20])
        _hc(ws,1,1,"SSE-element",18,"1B5E20")
        for ji,r in enumerate(sse_all,2): _hc(ws,1,ji,f"{r['freq']:.2f}",9,"1B5E20")
        for ri,sn in enumerate(sse_names,2):
            ws.cell(ri,1,sn).font=Font(name="Arial",bold=True,size=9)
            for ji,r in enumerate(sse_all,2):
                _dc(ws,ri,ji,r["sse"].get(sn,{}).get(metric,0.))


def _ws_sse_umap(wb, results, sse_umap_data, runlog):
    """Sheets 'SSE_UMAP_Cluster' and 'SSE_UMAP_Profil': Koordinaten + Z-Score-Profil."""
    try:
        Z2d, full_labels, feat_names, X_norm, valid_idx, cluster_chars = sse_umap_data
        if Z2d is None: return
        vi   = {gi: li for li, gi in enumerate(valid_idx)}
        fills = ["C8E6C9","A5D6A7","81C784","66BB6A","4CAF50",
                 "B2DFDB","DCEDC8","F0F4C3","E8F5E9","C8E6C9"]

        # ── Sheet 1: Koordinaten ─────────────────────────────────────
        ws = wb.create_sheet("SSE_UMAP_clusters")
        for ci, (t, w) in enumerate([("Frequency", 12), ("Type", 14),
                                      ("SSE-Cluster", 13),
                                      ("UMAP Dim1", 12), ("UMAP Dim2", 12)], 1):
            _hc(ws, 1, ci, t, w, "1B5E20")
        for ri, r in enumerate(results, 2):
            gi = ri - 2
            if gi in set(valid_idx):
                li = vi[gi]; k = int(full_labels[gi])
                fill = fills[k % len(fills)] if k >= 0 else "EEEEEE"
                _dc(ws, ri, 1, r["freq"], fill)
                _dc(ws, ri, 2, r["mode_type"], fill)
                _dc(ws, ri, 3, k if k >= 0 else "noise", fill)
                _dc(ws, ri, 4, round(float(Z2d[li, 0]), 6), fill)
                _dc(ws, ri, 5, round(float(Z2d[li, 1]), 6), fill)
            else:
                _dc(ws, ri, 1, r["freq"]); _dc(ws, ri, 2, r["mode_type"])

        # ── Sheet 2: Z-Score-Profil ──────────────────────────────────
        if not cluster_chars: return
        ws2 = wb.create_sheet("SSE_UMAP_profile")
        cluster_ids = sorted(k for k in cluster_chars.keys()
                             if isinstance(k, (int, np.integer)) and k >= 0)
        n_cl = len(cluster_ids)

        # Titel
        ws2.cell(1, 1, f"Cluster-Analyse: SSE_UMAP  ({n_cl} Cluster)")
        ws2.cell(1, 1).font = Font(name="Arial", bold=True, size=10,
                                    color="FFFFFF")
        ws2.cell(1, 1).fill = PatternFill("solid", fgColor="1B5E20")
        ws2.merge_cells(start_row=1, start_column=1,
                         end_row=1, end_column=2 + n_cl * 2)
        ws2.row_dimensions[1].height = 22

        # Z-Score-Abschnitt
        ws2.cell(3, 1, "A) Z-Score-Profil (SSE-Amplitudenfeatures)")
        ws2.cell(3, 1).font = Font(name="Arial", bold=True, size=9,
                                    color="1B5E20")
        ri = 4
        # Header
        _hc(ws2, ri, 1, "Feature", 28, "2E7D32")
        _hc(ws2, ri, 2, "Fisher-F", 12, "2E7D32")
        for ci, k in enumerate(cluster_ids):
            cc = cluster_chars[k]
            n_k = sum(1 for gl in full_labels if gl == k)
            _hc(ws2, ri, 3 + ci * 2,     f"C{k}(n={n_k})", 10, "2E7D32")
            _hc(ws2, ri, 3 + ci * 2 + 1, f"C{k} Z",        9, "388E3C")
        ws2.row_dimensions[ri].height = 24
        ri += 1

        # Features after Fisher-F absteigend sortiert
        # chars["_fisher"] = Fisher-F-values (global), chars[k]["means"/"z_scores"]
        ref_k    = cluster_ids[0]
        fisher_f = np.array(cluster_chars.get("_fisher", []))
        n_feats  = len(fisher_f)
        order    = np.argsort(fisher_f)[::-1] if n_feats > 0 else []

        for fi in order:
            if fi >= len(feat_names): continue
            fname = feat_names[fi]
            parts = fname.rsplit("_", 1)
            short = f"{parts[0][-14:]}_{parts[1]}" if len(parts) == 2 else fname[-22:]
            ws2.cell(ri, 1, short).font = Font(name="Arial", size=8)
            ff = float(fisher_f[fi]) if n_feats > 0 else 0.
            ws2.cell(ri, 2, round(ff, 2)).font = Font(name="Arial", size=8)
            ws2.cell(ri, 2).number_format = "0.0"
            for ci, k in enumerate(cluster_ids):
                cc = cluster_chars.get(k, {})
                mean_v = float(cc["means"][fi])    if "means"    in cc and n_feats > 0 else 0.
                z_v    = float(cc["z_scores"][fi]) if "z_scores" in cc and n_feats > 0 else 0.
                ws2.cell(ri, 3 + ci * 2,     round(mean_v, 4)).font = Font(name="Arial", size=8)
                ws2.cell(ri, 3 + ci * 2 + 1, round(z_v,    3)).font = Font(name="Arial", size=8, italic=True)
            ri += 1

        # Repraesentativste Moden
        ri += 1
        ws2.cell(ri, 1, "B) Repraesentativste modes (nächste to Clusterzentroid)")
        ws2.cell(ri, 1).font = Font(name="Arial", bold=True, size=9, color="1B5E20")
        ri += 1
        for ci, k in enumerate(cluster_ids):
            ws2.cell(ri, 1 + ci * 4, f"C{k}").font = Font(name="Arial", bold=True, size=9)
        ri += 1
        for ci, k in enumerate(cluster_ids):
            for col, h in enumerate(["Rank","Freq.","Type","Mode#"], 1 + ci * 4):
                _hc(ws2, ri, col, h, 8, "388E3C")
        ri += 1
        top_n = cluster_chars[ref_k].get("top_n_modes", [])
        for rank_i in range(len(top_n)):
            for ci, k in enumerate(cluster_ids):
                top = cluster_chars[k].get("top_n_modes", [])
                if rank_i < len(top):
                    m = top[rank_i]
                    base = 1 + ci * 4
                    ws2.cell(ri, base,     rank_i + 1).font = Font(name="Arial", size=8)
                    ws2.cell(ri, base + 1, round(m["freq"], 4)).font = Font(name="Arial", size=8)
                    ws2.cell(ri, base + 2, m["mode_type"]).font = Font(name="Arial", size=8)
                    ws2.cell(ri, base + 3, m["number"]).font = Font(name="Arial", size=8)
            ri += 1
        ws2.freeze_panes = "A5"

    except Exception as e:
        runlog.warn(f"SSE-UMAP-Sheet: {e}")


def _ws_ca_umap(wb, results, ca_umap_data, runlog):
    """Sheets 'Ca_UMAP_clusters' and 'Ca_UMAP_profile': UMAP coords + Z-score profile.

    Mirrors ``_ws_sse_umap`` but consumes the output of
    ``compute_ca_umap_cluster`` and uses a blue theme (vs green for SSE-UMAP)
    to keep the two embeddings visually distinct.

    New in v1.0.3.
    """
    try:
        Z2d, full_labels, feat_names, X_norm, valid_idx, cluster_chars = ca_umap_data
        if Z2d is None: return
        vi_set = set(valid_idx)
        fills = ["BBDEFB","90CAF9","64B5F6","42A5F5","2196F3",
                 "B3E5FC","E1F5FE","81D4FA","4FC3F7","29B6F6"]

        # ── Sheet 1: coordinates ─────────────────────────────────────────
        ws = wb.create_sheet("Ca_UMAP_clusters")
        for ci, (t, w) in enumerate([("Frequency", 12), ("Type", 14),
                                      ("Ca-Cluster", 13),
                                      ("UMAP Dim1", 12), ("UMAP Dim2", 12)], 1):
            _hc(ws, 1, ci, t, w, "0D47A1")
        for ri, r in enumerate(results, 2):
            gi = ri - 2
            if gi in vi_set:
                k = int(full_labels[gi])
                fill = fills[k % len(fills)] if k >= 0 else "EEEEEE"
                z1 = Z2d[gi, 0]; z2 = Z2d[gi, 1]
                # Z2d_full from compute_ca_umap_cluster can have NaN rows for
                # modes outside valid_idx; for modes inside valid_idx values
                # are finite.
                _dc(ws, ri, 1, r["freq"], fill)
                _dc(ws, ri, 2, r["mode_type"], fill)
                _dc(ws, ri, 3, k if k >= 0 else "noise", fill)
                _dc(ws, ri, 4, round(float(z1), 6) if np.isfinite(z1) else None, fill)
                _dc(ws, ri, 5, round(float(z2), 6) if np.isfinite(z2) else None, fill)
            else:
                _dc(ws, ri, 1, r["freq"])
                _dc(ws, ri, 2, r["mode_type"])

        # ── Sheet 2: Z-score profile ─────────────────────────────────────
        if not cluster_chars: return
        ws2 = wb.create_sheet("Ca_UMAP_profile")
        cluster_ids = sorted(k for k in cluster_chars.keys()
                             if isinstance(k, (int, np.integer)) and k >= 0)
        n_cl = len(cluster_ids)

        # Title
        ws2.cell(1, 1, f"Cluster analysis: Ca_UMAP  ({n_cl} clusters)")
        ws2.cell(1, 1).font = Font(name="Arial", bold=True, size=10,
                                    color="FFFFFF")
        ws2.cell(1, 1).fill = PatternFill("solid", fgColor="0D47A1")
        ws2.merge_cells(start_row=1, start_column=1,
                         end_row=1, end_column=2 + n_cl * 2)
        ws2.row_dimensions[1].height = 22

        # Z-score section
        ws2.cell(3, 1, "A) Z-score profile (C-alpha amplitude features)")
        ws2.cell(3, 1).font = Font(name="Arial", bold=True, size=9,
                                    color="0D47A1")
        ri = 4
        _hc(ws2, ri, 1, "Ca residue", 18, "1565C0")
        _hc(ws2, ri, 2, "Fisher-F",   12, "1565C0")
        for ci, k in enumerate(cluster_ids):
            cc = cluster_chars[k]
            n_k = sum(1 for gl in full_labels if gl == k)
            _hc(ws2, ri, 3 + ci * 2,     f"C{k}(n={n_k})", 10, "1565C0")
            _hc(ws2, ri, 3 + ci * 2 + 1, f"C{k} Z",        9, "1976D2")
        ws2.row_dimensions[ri].height = 24
        ri += 1

        # Features ordered by Fisher-F descending; only show top 30 Ca residues
        # to keep the profile sheet compact (the full data lives in Ca_amplitudes).
        ref_k    = cluster_ids[0]
        fisher_f = np.array(cluster_chars.get("_fisher", []))
        n_feats  = len(fisher_f)
        order    = np.argsort(fisher_f)[::-1] if n_feats > 0 else []
        top_show = 30
        for fi in order[:top_show]:
            if fi >= len(feat_names): continue
            fname = feat_names[fi]
            ws2.cell(ri, 1, fname).font = Font(name="Arial", size=8)
            ff = float(fisher_f[fi]) if n_feats > 0 else 0.
            ws2.cell(ri, 2, round(ff, 2)).font = Font(name="Arial", size=8)
            ws2.cell(ri, 2).number_format = "0.0"
            for ci, k in enumerate(cluster_ids):
                cc = cluster_chars.get(k, {})
                mean_v = float(cc["means"][fi])    if "means"    in cc and n_feats > 0 else 0.
                z_v    = float(cc["z_scores"][fi]) if "z_scores" in cc and n_feats > 0 else 0.
                ws2.cell(ri, 3 + ci * 2,     round(mean_v, 4)).font = Font(name="Arial", size=8)
                ws2.cell(ri, 3 + ci * 2 + 1, round(z_v,    3)).font = Font(name="Arial", size=8, italic=True)
            ri += 1

        # Representative modes per cluster (closest to centroid)
        ri += 1
        ws2.cell(ri, 1, "B) Representative modes (closest to cluster centroid)")
        ws2.cell(ri, 1).font = Font(name="Arial", bold=True, size=9, color="0D47A1")
        ri += 1
        for ci, k in enumerate(cluster_ids):
            ws2.cell(ri, 1 + ci * 4, f"C{k}").font = Font(name="Arial", bold=True, size=9)
        ri += 1
        for ci, k in enumerate(cluster_ids):
            for col, h in enumerate(["Rank", "Freq.", "Type", "Mode#"], 1 + ci * 4):
                _hc(ws2, ri, col, h, 8, "1976D2")
        ri += 1
        top_n = cluster_chars[ref_k].get("top_n_modes", [])
        for rank_i in range(len(top_n)):
            for ci, k in enumerate(cluster_ids):
                top = cluster_chars[k].get("top_n_modes", [])
                if rank_i < len(top):
                    m = top[rank_i]
                    base = 1 + ci * 4
                    ws2.cell(ri, base,     rank_i + 1).font = Font(name="Arial", size=8)
                    ws2.cell(ri, base + 1, round(m["freq"], 4)).font = Font(name="Arial", size=8)
                    ws2.cell(ri, base + 2, m["mode_type"]).font = Font(name="Arial", size=8)
                    ws2.cell(ri, base + 3, m["number"]).font = Font(name="Arial", size=8)
            ri += 1
        ws2.freeze_panes = "A5"

    except Exception as e:
        runlog.warn(f"Ca-UMAP sheet: {e}")


def _ws_ca(wb, results, ca_data, runlog):
    """Sheet 'Calpha_Amplituden': Calpha-displacementsamplituden per Residuum and Mode."""
    try:
        _,ca_res_nrs,ca_matrix=ca_data
        if ca_matrix is None: return
        ws=wb.create_sheet("Ca_amplitudes")
        _hc(ws,1,1,"Residue",12,"2E4057")
        for ji,r in enumerate(results,2): _hc(ws,1,ji,f"{r['freq']:.2f}",9,"2E4057")
        for ri,res_nr in enumerate(ca_res_nrs,2):
            ws.cell(ri,1,f"CA_{res_nr}").font=Font(name="Arial",size=9)
            for mi in range(len(results)):
                if ri-2<ca_matrix.shape[0] and mi<ca_matrix.shape[1]:
                    _dc(ws,ri,mi+2,float(ca_matrix[ri-2,mi]))
        ws.freeze_panes="B2"
    except Exception as e: runlog.warn(f"Ca-Sheet: {e}")


def _ws_projektionen(wb, results, embedding_coords):
    """Sheet 'Projektionen': PCA-, t-SNE- and UMAP-Koordinaten aller modes."""
    ws=wb.create_sheet("Projections")
    hdrs=[("Frequency",12),("Type",14),("Core mode",14)]
    for m in embedding_coords: hdrs+=[(f"{m}_Dim1",12),(f"{m}_Dim2",12)]
    for ci,(h,w) in enumerate(hdrs,1): _hc(ws,1,ci,h,w,"7D3C98")
    ws.row_dimensions[1].height=28
    for ri,r in enumerate(results,2):
        rf=_MTYPE_FILL.get(r["mode_type"]); idx=ri-2
        _dc(ws,ri,1,r["freq"],rf); _dc(ws,ri,2,r["mode_type"],rf)
        _dc(ws,ri,3,r.get("kern_primary","n/a"),rf)
        for ci,(m,Z) in enumerate(embedding_coords.items()):
            _dc(ws,ri,4+ci*2,round(float(Z[idx,0]),6) if idx<len(Z) else None,rf)
            _dc(ws,ri,5+ci*2,round(float(Z[idx,1]),6) if idx<len(Z) else None,rf)
    ws.freeze_panes="A2"


def _ws_cluster(wb, results, cluster_data, embedding_coords, feat_matrix, feat_names):
    """Sheet 'Cluster_<Methode>': Cluster-Labels and 2-D-Einbettung per Methode."""
    CL=["BBDEFB","FFCDD2","C8E6C9","FFE0B2","E1BEE7","B2EBF2","D7CCC8","FCE4EC"]
    for method,(labels,chars,cids) in cluster_data.items():
        try:
            Z2d=embedding_coords.get(method)
            if Z2d is None: continue
            ws=wb.create_sheet(f"{method[:12]}_Cluster")
            cl_hdrs=[("Frequency",12),("Type",14),("Cluster ID",10),("Distance",12)]
            cl_hdrs+=[(fn, 9) for fn in feat_names]
            for ci,(h,w) in enumerate(cl_hdrs,1): _hc(ws,1,ci,h,w,"1F4E79")
            centers2d={k:Z2d[labels==k].mean(0) for k in cids if (labels==k).any()}
            for ri,r in enumerate(results,2):
                idx=ri-2; k=int(labels[idx]) if idx<len(labels) else -1
                dist=(float(np.linalg.norm(Z2d[idx]-centers2d[k]))
                      if k>=0 and k in centers2d else 0.)
                fill=CL[k%len(CL)] if k>=0 else "EEEEEE"
                _dc(ws,ri,1,r["freq"],fill); _dc(ws,ri,2,r["mode_type"],fill)
                _dc(ws,ri,3,k if k>=0 else "noise",fill)
                _dc(ws,ri,4,round(dist,5),fill)
                if idx<feat_matrix.shape[0]:   # B4-Fix
                    for col_fi,val in enumerate(feat_matrix[idx],5):
                        _dc(ws,ri,col_fi,float(val),fill)
            ws.freeze_panes="A2"
        except Exception as _e:
            import warnings as _w; _w.warn(f"[export] Feature-Sheet Error: {_e}")


def _ws_cluster_profil(wb, results, cluster_data, feat_names):
    """Sheet 'Cluster_Profil_<Methode>': Top-modes and Feature-Profile jeof the cluster."""
    CL=["BBDEFB","C8E6C9","FFE0B2","F8BBD9","E1BEE7","B2DFDB"]
    for method,(labels,chars,cids) in cluster_data.items():
        try:
            ws=wb.create_sheet(f"{method[:12]}_Cluster_Profil")
            fisher=chars.get("_fisher",np.zeros(len(feat_names)))
            ws.cell(1,1,f"Cluster-Analyse: {method}").font=Font(name="Arial",bold=True,size=11)
            ws.cell(3,1,"A) Z-Score-Profil").font=Font(name="Arial",bold=True,size=9,color="1F4E79")
            _hc(ws,4,1,"Feature",26,"1F4E79"); _hc(ws,4,2,"Fisher-F",11,"1F4E79")
            co=3
            for k in cids:
                _hc(ws,4,co,f"C{k}(n={chars[k]['n']})",11,"1F4E79")
                _hc(ws,4,co+1,f"C{k} Z",10,"1F4E79"); co+=2
            ws.row_dimensions[4].height=30
            for row,fi in enumerate(np.argsort(-fisher).tolist(),5):
                ws.cell(row,1,feat_names[fi]).font=Font(name="Arial",size=9)
                fv=float(fisher[fi])
                ws.cell(row,2,round(fv,4) if fv>0 else None).font=Font(name="Arial",size=9)
                co=3
                for k in cids:
                    z=float(chars[k]["z_scores"][fi])
                    _dc(ws,row,co,round(float(chars[k]["means"][fi]),4))
                    cz=ws.cell(row,co+1,round(z,3)); cz.font=Font(name="Arial",size=9)
                    if z<-1.5: cz.fill=PatternFill("solid",fgColor="FFB3B3")
                    elif z>1.5: cz.fill=PatternFill("solid",fgColor="B3FFB3")
                    co+=2
            B=len(feat_names)+7
            ws.cell(B-1,1,"B) Repraesentativste Moden").font=\
                Font(name="Arial",bold=True,size=9,color="1F4E79")
            bc=1
            for k in cids:
                _hc(ws,B,bc,f"C{k}(n={chars[k]['n']})",10,"1F4E79")
                _hc(ws,B+1,bc,"Rank",7,"1F4E79"); _hc(ws,B+1,bc+1,"Freq.",12,"1F4E79")
                _hc(ws,B+1,bc+2,"Type",14,"1F4E79"); _hc(ws,B+1,bc+3,"Mode#",9,"1F4E79")
                fl=CL[k%len(CL)]
                for ri,m in enumerate(chars[k]["top_n_modes"],B+2):
                    _dc(ws,ri,bc,ri-B-1,fl); _dc(ws,ri,bc+1,round(m["freq"],4),fl)
                    _dc(ws,ri,bc+2,m["mode_type"],fl); _dc(ws,ri,bc+3,m["number"],fl)
                bc+=5
        except Exception as _e:
            import warnings as _w; _w.warn(f"[export] Cluster-Profil Error: {_e}")


__version__ = "1.4"  # modenanalyse v1.4


# ===========================================================================
# v3.7 Sheets: Marcus-Hush-reorganization energyn
# ===========================================================================

def _ws_reorganisationsenergie_v37(wb, results):
    """Sheet 'reorganization energy': Pro-Mode dr_X and lambda_X fuer
    alle bond channels.

    Origin-freundlich: reine Datentabelle, no Klassifikation,
    no Faerbung in the Datenbereich. A Zeile per Mode, vier Spalten
    per Kanal (dr_rss, dr_signed_summe, lambda_pair, lambda_mode).
    """
    ws = wb.create_sheet("Reorganization_energy")
    
    from .reorganization import CHANNELS
    
    # Header: freq, dann per Kanal X the 4 Spalten
    headers = ["freq_cm1"]
    for ch in CHANNELS:
        headers.append(f"dr_{ch}_rss_a")
        headers.append(f"dr_{ch}_sum_signed_a")
        headers.append(f"lambda_{ch}_pair_cm1")
        headers.append(f"lambda_{ch}_mode_cm1")
    
    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, h)
        c.font = Font(name="Arial", bold=True, size=10)
        ws.column_dimensions[get_column_letter(ci)].width = 18
    
    rows = sorted(results, key=lambda r: r.get("freq", 0.0))
    
    def _v(x):
        if x is None: return None
        if isinstance(x, float) and not np.isfinite(x): return None
        if isinstance(x, (int, float)): return float(x)
        return None
    
    for ri, r in enumerate(rows, 2):
        agg = r.get("reorg_per_mode", {}) or {}
        values = [_v(r.get("freq"))]
        for ch in CHANNELS:
            d = agg.get(ch, {}) or {}
            values.append(_v(d.get("dr_rss_a")))
            values.append(_v(d.get("dr_sum_signed_a")))
            values.append(_v(d.get("lambda_pair_cm1")))
            values.append(_v(d.get("lambda_mode_cm1")))
        for ci, v in enumerate(values, 1):
            c = ws.cell(ri, ci)
            if v is not None:
                c.value = round(v, 6) if isinstance(v, float) else v
            c.font = Font(name="Arial", size=9)
    
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _ws_reorg_pro_bindung_v37(wb, aggregates):
    """Sheet 'Reorg_pro_bond': per einzelner bond (Sub-Channel)
    seine eigene Marcus-Hush-Reorg.

    Im Gegensatz to Reorg_Total, the per Parent-Channel (FeFe, FeN, FeS,
    NH, HA) aggregiert, zeigt dieses Sheet jede einzelne bond
    separat: FeS_Cys207, FeS_Cys216, FeS_Cluster_2_3, FeN_His255,
    FeN_His259, HA_His255_O6265, etc. So sieht man, welche einzelne
    bond at the meisten to Reorganisation beicontributes.
    """
    ws = wb.create_sheet("Reorg_per_bond")
    
    sub_totals = aggregates.get("sub_totals", {}) or {}
    
    headers = ["bond", "Parent_Kanal", "acceptor_Gewicht",
               "Lambda_total_pair_cm1", "Lambda_total_mode_cm1",
               "n_modes_contributing"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, h)
        c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0D47A1")
        ws.column_dimensions[get_column_letter(ci)].width = 24
    
    # Sortieren after Parent-Channel, dann after Lambda absteigend
    items = list(sub_totals.items())
    items.sort(key=lambda kv: (
        kv[1].get("parent_channel", ""),
        -kv[1].get("lambda_total_mode_cm1", 0.0)
    ))
    
    for ri, (name, d) in enumerate(items, 2):
        ws.cell(ri, 1, name).font = Font(name="Arial", size=10)
        ws.cell(ri, 2, d.get("parent_channel", "")).font = Font(name="Arial", size=10)
        w = d.get("weight", 1.0)
        ws.cell(ri, 3, round(w, 4)).font = Font(name="Arial", size=10)
        ws.cell(ri, 4, round(d.get("lambda_total_pair_cm1", 0.0), 4)).font = Font(name="Arial", size=10)
        ws.cell(ri, 5, round(d.get("lambda_total_mode_cm1", 0.0), 4)).font = Font(name="Arial", size=10)
        ws.cell(ri, 6, d.get("n_modes_contributing", 0)).font = Font(name="Arial", size=10)
    
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _ws_reorg_total_v37(wb, aggregates):
    """Sheet 'Reorg_Total': system total Marcus-Hush-Reorg per Kanal.

    A kompakte Tabelle: per Kanal the Total-Reorg (Lambda_total)
    in cm-1, sowohl with Mode-mu als also with Pair-mu. A Zeile pro
    Kanal, plus Header. Diese is the zentrale System-Charakterisierung.
    """
    ws = wb.create_sheet("Reorg_total")
    
    headers = ["Channel", "Lambda_total_pair_cm1", "Lambda_total_mode_cm1",
               "n_modes_contributing"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, h)
        c.font = Font(name="Arial", bold=True, size=11)
        c.fill = PatternFill("solid", fgColor="0D47A1")
        c.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        ws.column_dimensions[get_column_letter(ci)].width = 24
    
    totals = aggregates.get("totals", {}) or {}
    from .reorganization import CHANNELS
    
    for ri, ch in enumerate(CHANNELS, 2):
        t = totals.get(ch, {}) or {}
        ws.cell(ri, 1, ch).font = Font(name="Arial", bold=True, size=10)
        for ci, key in enumerate(["lambda_total_pair_cm1",
                                    "lambda_total_mode_cm1",
                                    "n_modes_contributing"], 2):
            v = t.get(key, 0.0)
            if isinstance(v, float):
                v = round(v, 4)
            ws.cell(ri, ci, v).font = Font(name="Arial", size=10)
    
    ws.freeze_panes = "A2"


def _ws_modulations_spektren_v37(wb, aggregates):
    """Sheet 'Modulations_Spektren': frequenz-aufgelostes M_X(omega).

    Spalten:
      freq_cm1, M_FeFe, M_FeN, M_FeS, M_NH, M_HA,
      C_PCET (=sqrt(M_HA*M_FeFe)), C_PT_FeN (=sqrt(M_HA*M_FeN)),
      C_ET_FeS (=sqrt(M_FeFe*M_FeS))

    Origin-freundlich: per Spalte a Kurve. Direkt with NRVS-Spektren
    ueberlagerbar.
    """
    ws = wb.create_sheet("Modulation_spectra")
    
    grid = aggregates.get("spectra_grid_cm1")
    spectra = aggregates.get("modulation_spectra", {}) or {}
    co_spectra = aggregates.get("co_modulation_spectra", {}) or {}
    sigma = aggregates.get("spectra_sigma_cm1", 5.0)
    
    if grid is None or len(grid) == 0:
        ws.cell(1, 1, "(keine Spektren-Daten)")
        return
    
    from .reorganization import CHANNELS
    
    # Header in Zeile 1, Hinweis in Zeile 2
    headers = ["freq_cm1"]
    headers += [f"M_{ch}" for ch in CHANNELS]
    headers += sorted(co_spectra.keys())
    
    note = (f"Modulations-Spektren M_X(omega) = sum_i |dr_X(i)| * "
            f"Gauss(omega-omega_i, sigma={sigma:.1f} cm-1). "
            f"Co-Spektren = geometr. Mittel zweier M_X.")
    ws.cell(1, 1, note).font = Font(name="Arial", italic=True, size=8,
                                     color="666666")
    
    for ci, h in enumerate(headers, 1):
        c = ws.cell(2, ci, h)
        c.font = Font(name="Arial", bold=True, size=10)
        c.fill = PatternFill("solid", fgColor="0D47A1")
        c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        ws.column_dimensions[get_column_letter(ci)].width = 14
    
    for ri, f in enumerate(grid, 3):
        ws.cell(ri, 1, float(round(f, 4))).font = Font(name="Arial", size=8)
        col = 2
        for ch in CHANNELS:
            arr = spectra.get(ch)
            v = float(arr[ri-3]) if arr is not None and ri-3 < len(arr) else 0.0
            if v != 0.0:
                ws.cell(ri, col, round(v, 6)).font = Font(name="Arial", size=8)
            col += 1
        for k in sorted(co_spectra.keys()):
            arr = co_spectra.get(k)
            v = float(arr[ri-3]) if arr is not None and ri-3 < len(arr) else 0.0
            if v != 0.0:
                ws.cell(ri, col, round(v, 6)).font = Font(name="Arial", size=8)
            col += 1
    
    ws.freeze_panes = "B3"


def _ws_lambda_kumulativ_v37(wb, aggregates):
    """Sheet 'Lambda_kumulativ': kumulative Reorg-Kurven Lambda_X(omega).

    Spalten:
      freq_cm1, Lambda_FeFe, Lambda_FeN, Lambda_FeS, Lambda_NH, Lambda_HA

    Diese are monoton steigend; konvergieren at the Spektrum-Ende gegen
    the Total-value from the Reorg_Total-Sheet.
    """
    ws = wb.create_sheet("Lambda_cumulative")
    
    grid = aggregates.get("spectra_grid_cm1")
    cum = aggregates.get("cumulative_reorg", {}) or {}
    use_mode = aggregates.get("cumulative_uses_mode_mass", True)
    
    if grid is None or len(grid) == 0:
        ws.cell(1, 1, "(keine kumulativen Daten)")
        return
    
    from .reorganization import CHANNELS
    
    note = (f"Lambda_X(omega) = sum_{{i: omega_i <= omega}} lambda_X(i) "
            f"in cm-1. {'Mode-mu' if use_mode else 'Pair-mu'}-Konvention.")
    ws.cell(1, 1, note).font = Font(name="Arial", italic=True, size=8,
                                     color="666666")
    
    headers = ["freq_cm1"] + [f"Lambda_{ch}" for ch in CHANNELS]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(2, ci, h)
        c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0D47A1")
        ws.column_dimensions[get_column_letter(ci)].width = 14
    
    for ri, f in enumerate(grid, 3):
        ws.cell(ri, 1, float(round(f, 4))).font = Font(name="Arial", size=8)
        for ci, ch in enumerate(CHANNELS, 2):
            arr = cum.get(ch)
            v = float(arr[ri-3]) if arr is not None and ri-3 < len(arr) else 0.0
            if v != 0.0:
                ws.cell(ri, ci, round(v, 6)).font = Font(name="Arial", size=8)
    
    ws.freeze_panes = "B3"

