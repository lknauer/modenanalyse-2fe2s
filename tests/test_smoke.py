# Copyright (C) 2026 Lukas Knauer, AG Schuenemann, RPTU Kaiserslautern-Landau
# SPDX-License-Identifier: GPL-3.0-or-later
"""End-to-End-Smoke-Test fuer modenanalyse_2fe2s.

Nutzt das Cys-2Fe-2S-Modell-Logfile (Cys4-Cluster, 52 Atome, 150 Modes,
~1 MB komprimiert auf ~150 KB) als realistisches Mini-System. Testet die
gesamte Pipeline:

  Logfile -> Cluster-Erkennung -> Reorg-Aggregation -> Excel-Output

``run_analysis`` wird nur einmal pro Test-Modul aufgerufen
(``scope="module"`` Fixture); die einzelnen Tests pruefen dann
verschiedene Aspekte des Outputs.

Marker: ``slow`` (pytest -m slow). Default ist diese Suite ausgeschlossen,
weil sie ~60 Sekunden braucht (volle Pipeline). Aufruf:

  pytest tests/test_smoke.py -m slow -v

Was getestet wird:

* Tool laeuft fehlerfrei auf einem realen Logfile
* Cluster-Erkennung liefert plausible Geometrie
* Output-Dateien werden geschrieben (analysis.xlsx, REPORT.txt, etc.)
* "no-His"-Pfad: PCET-Reorg wird sauber als nicht-aktiv erkannt (Cys4)
* Reorg-Werte sind finite und im erwarteten Groessenbereich
* Marcus-Hush-Additivitaet: Total = Summe der pro-Mode-Beitraege
"""

from __future__ import annotations
import lzma
import shutil
import tempfile
from pathlib import Path

import numpy as np
import pytest


# Mark this whole module as "slow" so it's skipped by default.
pytestmark = pytest.mark.slow


# Path to the compressed test logfile (relative to this test file).
DATA_DIR = Path(__file__).parent / "data"
LOG_XZ = DATA_DIR / "Cys_2Fe-2S_red3_hpfrq_opt.log.xz"


@pytest.fixture(scope="module")
def smoke_run_outputs(tmp_path_factory):
    """Fuehrt einen vollstaendigen run_analysis-Lauf einmalig aus.
    
    Die einzelnen Tests teilen sich diese Outputs (scope='module'),
    sodass das ~50-Sekunden-Setup nur einmal noetig ist statt pro Test.
    
    Liefert ein Dict mit:
      - 'output_dir' (Path): Wurzel-Verzeichnis der Outputs
      - 'sub_dir'    (Path): Unter-Verzeichnis "0-800_cm-1"
      - 'befund'     (str):  Inhalt der REPORT.txt
      - 'excel'      (Path): Pfad zur _analysis.xlsx
    """
    from modenanalyse_2fe2s import Config, run_analysis
    
    assert LOG_XZ.exists(), f"Test-Datei fehlt: {LOG_XZ}"
    
    # Permanenter Tempdir fuer die Modul-Laufzeit.
    workdir = tmp_path_factory.mktemp("smoke_run")
    log_path = workdir / "Cys_2Fe-2S_red3_hpfrq_opt.log"
    
    # Logfile entpacken
    with lzma.open(LOG_XZ, "rb") as src:
        with open(log_path, "wb") as dst:
            shutil.copyfileobj(src, dst)
    
    output_dir = workdir / "results"
    cfg = Config(
        log_file = str(log_path),
        output_dir = str(output_dir),
        pdb_file = "",
        temp_k = 5.0,
        freq_max = 800.0,
        analyze_scsd = True,
        pcet_enabled = True,
        use_cache = False,
    )
    run_analysis(cfg)
    
    sub_dir = output_dir / "0-800_cm-1"
    befund_files = list(sub_dir.glob("*_REPORT.txt"))
    excel_files = list(sub_dir.glob("*_analysis.xlsx"))
    assert befund_files, f"REPORT.txt nicht gefunden in {sub_dir}"
    assert excel_files, f"_analysis.xlsx nicht gefunden in {sub_dir}"
    
    return {
        "output_dir": output_dir,
        "sub_dir": sub_dir,
        "befund": befund_files[0].read_text(encoding="utf-8", errors="replace"),
        "excel": excel_files[0],
    }


# ---------------------------------------------------------------------------
# Test 1: Output-Struktur
# ---------------------------------------------------------------------------

def test_smoke_output_files_exist(smoke_run_outputs):
    """Pipeline produziert die erwarteten Output-Dateien."""
    sub_dir = smoke_run_outputs["sub_dir"]
    
    excel_files = list(sub_dir.glob("*_analysis.xlsx"))
    assert len(excel_files) == 1, \
        f"Erwartet genau 1 *_analysis.xlsx, gefunden: {excel_files}"
    
    befund_files = list(sub_dir.glob("*_REPORT.txt"))
    assert len(befund_files) == 1, \
        f"Erwartet genau 1 *_REPORT.txt, gefunden: {befund_files}"
    
    embedding_xlsx = list(sub_dir.glob("*_Embeddings.xlsx"))
    assert len(embedding_xlsx) == 1, \
        f"Erwartet genau 1 *_Embeddings.xlsx, gefunden: {embedding_xlsx}"
    
    umap_png = list(sub_dir.glob("*embedding_UMAP.png"))
    assert len(umap_png) == 1, \
        f"Erwartet genau 1 UMAP-PNG, gefunden: {umap_png}"


# ---------------------------------------------------------------------------
# Test 2: REPORT-Inhalt
# ---------------------------------------------------------------------------

def test_smoke_befund_content(smoke_run_outputs):
    """REPORT file contains the most important expected markers.
    
    - PCET reorg marked as 'NOT active' (Cys4 cluster, no His)
    - Cluster geometry detected (Fe-Fe distance, Fe-S mean)
    - Reorg total values for FeFe and FeS reported
    """
    befund = smoke_run_outputs["befund"]
    
    # PCET reorg must be explicitly marked as not-active
    assert "PCET reorg: NOT active (no His" in befund, \
        "Expected: 'PCET reorg: NOT active (no His' in REPORT"
    
    # Cluster geometry
    assert "Fe-Fe=" in befund, "Expected: Fe-Fe distance in REPORT"
    assert "Fe-S(mean)=" in befund, "Expected: Fe-S mean in REPORT"
    
    # Reorg values
    assert "Lambda_FeFe=" in befund, "Expected: Lambda_FeFe in REPORT"
    assert "Lambda_FeS=" in befund, "Expected: Lambda_FeS in REPORT"


# ---------------------------------------------------------------------------
# Test 3: Reorg-Werte plausibel
# ---------------------------------------------------------------------------

def test_smoke_reorg_values_plausible(smoke_run_outputs):
    """Reorg-Werte sind physikalisch plausibel.
    
    - FeFe und FeS finit + positiv
    - FeS dominiert ueber FeFe (Cluster-Atmung << Fe-S-Streckung)
    - FeN, NH, HA = 0 (kein His-N, kein H-Bridge)
    - Sanity-Range fuer Cys4-Modell-Cluster
    """
    import openpyxl
    
    wb = openpyxl.load_workbook(smoke_run_outputs["excel"],
                                  read_only=True, data_only=True)
    assert "Reorg_total" in wb.sheetnames, \
        f"Sheet 'Reorg_total' missing; available: {wb.sheetnames}"
    
    ws = wb["Reorg_total"]
    rows = list(ws.values)
    
    # Layout: Header in Z. 1, eine Zeile pro Kanal
    # Spalten: Kanal, Lambda_total_pair_cm1, Lambda_total_mode_cm1, n_modes
    vals = {}
    for row in rows[1:]:
        if not row or row[0] is None:
            continue
        ch = str(row[0]).strip()
        if ch in ("FeFe", "FeN", "FeS", "NH", "HA"):
            vals[ch] = {
                "pair": float(row[1]) if row[1] is not None else 0.0,
                "mode": float(row[2]) if row[2] is not None else 0.0,
                "n":    int(row[3])   if row[3] is not None else 0,
            }
    
    # FeFe und FeS positiv und finite
    for ch in ("FeFe", "FeS"):
        assert vals[ch]["mode"] > 0, \
            f"{ch} Lambda_mode muss > 0 sein: {vals[ch]}"
        assert np.isfinite(vals[ch]["mode"]), \
            f"{ch} Lambda_mode nicht finite"
    
    # Hierarchie: FeS > FeFe (4 Fe-S-Bindungen vs 1 Fe-Fe)
    assert vals["FeS"]["mode"] > vals["FeFe"]["mode"], \
        f"Erwartet FeS > FeFe, bekommen FeS={vals['FeS']['mode']}, " \
        f"FeFe={vals['FeFe']['mode']}"
    
    # His-abhaengige Kanaele: alle null
    for ch in ("FeN", "NH", "HA"):
        assert vals[ch]["mode"] == 0, \
            f"{ch} muss 0 sein (kein His): {vals[ch]}"
    
    # Sanity-Range fuer Cys4-Modell-Cluster
    assert 1 < vals["FeFe"]["mode"] < 100, \
        f"FeFe ausserhalb Sanity-Range 1-100 cm-1: {vals['FeFe']['mode']}"
    assert 50 < vals["FeS"]["mode"] < 1000, \
        f"FeS ausserhalb Sanity-Range 50-1000 cm-1: {vals['FeS']['mode']}"
    
    # Mode-Counts konsistent
    assert vals["FeFe"]["n"] == vals["FeS"]["n"], \
        "n_modes_contributing inkonsistent zwischen FeFe und FeS"


# ---------------------------------------------------------------------------
# Test 4: Marcus-Hush-Additivitaet
# ---------------------------------------------------------------------------

def test_smoke_marcus_hush_additivity(smoke_run_outputs):
    """Lambda_total_mode = Summe der pro-Mode-Beitraege (Marcus-Hush).
    
    Diese Eigenschaft ist die zentrale Konsistenz-Bedingung der
    Implementation und wurde in einem fruehen Validierungslauf bereits
    auf 0.01 cm^-1 verifiziert.
    """
    import openpyxl
    
    wb = openpyxl.load_workbook(smoke_run_outputs["excel"],
                                  read_only=True, data_only=True)
    
    # Per-mode contributions from 'Reorganization_energy' sheet
    ws = wb["Reorganization_energy"]
    rows = list(ws.values)
    headers = rows[0]
    idx_fefe = headers.index("lambda_FeFe_mode_cm1")
    idx_fes  = headers.index("lambda_FeS_mode_cm1")
    
    sum_fefe = 0.0
    sum_fes  = 0.0
    for row in rows[1:]:
        v = row[idx_fefe]
        if v is not None and np.isfinite(float(v)):
            sum_fefe += float(v)
        v = row[idx_fes]
        if v is not None and np.isfinite(float(v)):
            sum_fes += float(v)
    
    # Total values from 'Reorg_total' sheet
    ws_total = wb["Reorg_total"]
    rows_total = list(ws_total.values)
    totals = {}
    for row in rows_total[1:]:
        if not row or row[0] is None:
            continue
        ch = str(row[0]).strip()
        if ch in ("FeFe", "FeS"):
            totals[ch] = float(row[2])  # lambda_total_mode_cm1
    
    # Vergleich: pro-Mode-Summe = Total (Toleranz 0.01 cm-1)
    assert abs(sum_fefe - totals["FeFe"]) < 0.01, \
        f"FeFe-Additivitaet verletzt: sum={sum_fefe:.4f}, " \
        f"total={totals['FeFe']:.4f}, diff={sum_fefe - totals['FeFe']:.4f}"
    
    assert abs(sum_fes - totals["FeS"]) < 0.01, \
        f"FeS-Additivitaet verletzt: sum={sum_fes:.4f}, " \
        f"total={totals['FeS']:.4f}, diff={sum_fes - totals['FeS']:.4f}"
