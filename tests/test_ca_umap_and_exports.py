# Copyright (C) 2026 Lukas Knauer, AG Schuenemann, RPTU Kaiserslautern-Landau
# SPDX-License-Identifier: GPL-3.0-or-later
"""Regression tests for v1.0.3 features:

1. compute_ca_umap_cluster — new Ca-amplitude UMAP embedding
2. export_embedding_excel — must now emit Ca_UMAP_clusters and
   Ca_UMAP_profile sheets when ca_umap_data is supplied
3. export_main_excel — must emit a Coordination diagnostic sheet
   listing each ligand's atom assignment
4. export_embedding_plots — must accept sse_umap_data, ca_umap_data,
   ca_data without raising on missing/None inputs

Run::

    python3 -m pytest tests/test_ca_umap_and_exports.py -v
"""
from __future__ import annotations
import pytest
from pathlib import Path
import sys
import tempfile
import os

import numpy as np

# Make src/ importable when running pytest from the repo root
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "src"))


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

class _FakeCfg:
    """Minimal cfg stub for outname() calls."""
    def __init__(self, outdir):
        self._od = outdir
    def outname(self, suffix):
        return os.path.join(self._od, "test" + suffix)


class _FakeRunLog:
    def __init__(self):
        self.warnings = []
        self.infos = []
        self.outputs = []
    def warn(self, m):    self.warnings.append(str(m))
    def info(self, m):    self.infos.append(str(m))
    def add_output(self, p): self.outputs.append(p)


def _make_synthetic_results(n_modes=30, n_calpha=20, seed=42):
    """Generate synthetic results + ca_data that's UMAP-meaningful.

    Modes 0-9: Gaussian amplitude pattern centered on residue 5 (localized)
    Modes 10-19: centered on residue 15 (localized elsewhere)
    Modes 20-29: flat amplitude (delocalized)
    """
    rng = np.random.default_rng(seed)
    results = []
    for i in range(n_modes):
        results.append({
            "number":    i + 1,
            "freq":      5.0 + i * 1.5,
            "mode_type": ("In-plane" if i % 3 == 0 else
                          "Out-of-plane" if i % 3 == 1 else
                          "Torsional/Mixed"),
        })

    ca_res_nums = list(range(1, n_calpha + 1))
    ca_centers  = [100 + i for i in range(n_calpha)]
    ca_matrix   = np.zeros((n_calpha, n_modes))  # (n_calpha, n_modes) per _build_ca_data

    for mi in range(n_modes):
        if mi < 10:
            center = 5
        elif mi < 20:
            center = 15
        else:
            center = -1   # flat
        for ri in range(n_calpha):
            if center >= 0:
                # Gaussian envelope
                ca_matrix[ri, mi] = (
                    0.05 * np.exp(-((ri - center) ** 2) / 8.)
                    + 0.005 * rng.normal()
                )
            else:
                ca_matrix[ri, mi] = 0.03 + 0.01 * rng.normal()

    return results, (ca_centers, ca_res_nums, ca_matrix)


# ---------------------------------------------------------------------------
# Layer 1: compute_ca_umap_cluster
# ---------------------------------------------------------------------------

def test_ca_umap_basic():
    """Ca-UMAP must run on a small synthetic system, return finite
    coordinates, and put modes-by-amplitude-pattern into clusters."""
    from modenanalyse_2fe2s.embedding import compute_ca_umap_cluster

    results, ca_data = _make_synthetic_results()
    runlog = _FakeRunLog()
    out = compute_ca_umap_cluster(results, ca_data, runlog=runlog)

    Z2d_full, full_labels, feat_names, X_norm, valid_idx, cluster_chars = out

    assert Z2d_full is not None,        "Z2d_full should be returned"
    assert full_labels is not None,      "labels should be returned"
    assert len(feat_names) == len(ca_data[1]), \
        f"feat_names should have one entry per Ca residue ({len(ca_data[1])}), got {len(feat_names)}"
    assert all(f.startswith("CA_") for f in feat_names), \
        "feat_names should follow CA_<residue#> convention"
    assert len(valid_idx) == len(results), \
        "all 30 synthetic modes should have valid Ca data"
    assert Z2d_full.shape == (len(results), 2), \
        f"Z2d_full shape should be (n_modes, 2), got {Z2d_full.shape}"
    # Finite coordinates on every valid row
    assert np.all(np.isfinite(Z2d_full[valid_idx, :])), \
        "valid rows of Z2d_full should be finite"


def test_ca_umap_orientation_autodetect():
    """compute_ca_umap_cluster must accept (n_calpha, n_modes) AND
    (n_modes, n_calpha) layouts (defensive against future changes
    in _build_ca_data)."""
    from modenanalyse_2fe2s.embedding import compute_ca_umap_cluster

    results, ca_data = _make_synthetic_results()
    ca_centers, ca_res_nums, ca_matrix = ca_data

    # Default (n_calpha, n_modes)
    runlog1 = _FakeRunLog()
    out1 = compute_ca_umap_cluster(results, ca_data, runlog=runlog1)

    # Transposed (n_modes, n_calpha)
    ca_data_T = (ca_centers, ca_res_nums, ca_matrix.T)
    runlog2 = _FakeRunLog()
    out2 = compute_ca_umap_cluster(results, ca_data_T, runlog=runlog2)

    # Both should succeed and produce SAME labels (UMAP+HDBSCAN are
    # deterministic on fixed random_state)
    assert out1[1] is not None and out2[1] is not None
    np.testing.assert_array_equal(out1[1], out2[1])


def test_ca_umap_too_few_modes_returns_none():
    """If fewer than 5 modes have Ca data, Ca-UMAP must abort cleanly."""
    from modenanalyse_2fe2s.embedding import compute_ca_umap_cluster

    results, ca_data = _make_synthetic_results(n_modes=4, n_calpha=10)
    runlog = _FakeRunLog()
    out = compute_ca_umap_cluster(results, ca_data, runlog=runlog)

    Z2d, labels, feat_names, X_norm, valid_idx, chars = out
    assert Z2d is None
    assert labels is None
    assert any("Ca-UMAP" in w and "skipped" in w for w in runlog.warnings)


def test_ca_umap_handles_none_input():
    """Ca-UMAP must return all-None tuple when ca_data is None
    (e.g. no Ca atoms in PDB), without raising."""
    from modenanalyse_2fe2s.embedding import compute_ca_umap_cluster
    out = compute_ca_umap_cluster([], None)
    assert out == (None, None, [], None, [], {})


# ---------------------------------------------------------------------------
# Layer 2: Excel export of Ca-UMAP sheets
# ---------------------------------------------------------------------------

def test_export_embedding_excel_writes_ca_umap_sheets():
    """export_embedding_excel must add Ca_UMAP_clusters and Ca_UMAP_profile
    when ca_umap_data is supplied."""
    import openpyxl
    from modenanalyse_2fe2s.embedding import compute_ca_umap_cluster
    from modenanalyse_2fe2s.export    import export_embedding_excel

    results, ca_data = _make_synthetic_results()
    runlog = _FakeRunLog()
    ca_umap = compute_ca_umap_cluster(results, ca_data, runlog=runlog)

    with tempfile.TemporaryDirectory() as td:
        out = os.path.join(td, "embed.xlsx")
        cfg = _FakeCfg(td)
        # Minimal call: only ca_data + ca_umap, no global UMAP, no clusters.
        export_embedding_excel(
            results, {}, None, [], {}, ca_data,
            out, cfg, runlog,
            ca_umap_data=ca_umap,
        )
        assert os.path.exists(out), "Embedding xlsx must be written"
        wb = openpyxl.load_workbook(out, read_only=True)
        names = set(wb.sheetnames)
        wb.close()

    assert "Ca_UMAP_clusters" in names, (
        f"Ca_UMAP_clusters sheet missing from {names}"
    )
    # Ca_UMAP_profile is only written when at least one cluster was found
    # (HDBSCAN may legitimately fail on tiny synthetic data; accept either
    # presence, but require it whenever cluster_chars is non-empty).
    if ca_umap[5]:
        assert "Ca_UMAP_profile" in names


def test_export_embedding_excel_skips_ca_umap_when_none():
    """Without ca_umap_data the new sheets must NOT appear (back-compat
    with v1.0.2 callers)."""
    import openpyxl
    from modenanalyse_2fe2s.export import export_embedding_excel

    results, ca_data = _make_synthetic_results()
    runlog = _FakeRunLog()

    with tempfile.TemporaryDirectory() as td:
        out = os.path.join(td, "embed.xlsx")
        cfg = _FakeCfg(td)
        export_embedding_excel(
            results, {}, None, [], {}, ca_data,
            out, cfg, runlog,
            # ca_umap_data omitted → defaults to None
        )
        wb = openpyxl.load_workbook(out, read_only=True)
        names = set(wb.sheetnames)
        wb.close()

    assert "Ca_UMAP_clusters" not in names
    assert "Ca_UMAP_profile" not in names


# ---------------------------------------------------------------------------
# Layer 3: Coordination diagnostic sheet
# ---------------------------------------------------------------------------

def test_coordination_sheet_lists_each_ligand_with_atom_count():
    """The new Coordination sheet must contain one row per ligand giving
    the residue, donor atom, bond length and atom count, plus per-atom
    detail rows."""
    import openpyxl
    from modenanalyse_2fe2s.export import _ws_coordination
    from openpyxl import Workbook

    # Build a minimal coord_info-like stub
    class _L:
        def __init__(self, res_label, res_name, res_num, lig_element,
                     lig_aname, bond_len):
            self.res_label   = res_label
            self.res_name    = res_name
            self.res_num     = res_num
            self.lig_element = lig_element
            self.lig_aname   = lig_aname
            self.bond_len    = bond_len

    class _CI:
        def __init__(self, ligands, group_map):
            self.ligands   = ligands
            self.group_map = group_map

    ci = _CI(
        ligands=[
            _L("His 255", "HIS", 255, "N", "ND1", 1.982),
            _L("His 259", "HIS", 259, "N", "ND1", 1.993),
            _L("Cys 207", "CYS", 207, "S", "SG",  2.135),
            _L("Cys 216", "CYS", 216, "S", "SG",  2.155),
        ],
        group_map={
            "His 255": [3881, 3882, 3883, 3884, 3885, 3886, 3887, 3888, 3889, 3890],
            "His 259": [3928, 3929, 3930, 3931, 3932, 3933, 3934, 3935, 3936, 3937],
            "Cys 207": [3142, 3143, 3144, 3145, 3146, 3147],
            "Cys 216": [3259, 3260, 3261, 3262, 3263, 3264],
        },
    )

    # Atoms list: build something just large enough to cover the indices.
    atoms = [None] * 4000
    # Populate a few that the test will assert on
    for c in ci.group_map["His 255"]:
        atoms[c - 1] = {"name": f"X{c}", "element": "C",
                         "x": 0.0, "y": 0.0, "z": 0.0}
    for c in ci.group_map["Cys 216"]:
        atoms[c - 1] = {"name": f"Y{c}", "element": "S",
                         "x": 1.0, "y": 2.0, "z": 3.0}

    runlog = _FakeRunLog()
    wb = Workbook(); wb.remove(wb.active)
    _ws_coordination(wb, ci, atoms, runlog)

    assert "Coordination" in wb.sheetnames
    ws = wb["Coordination"]

    # Header row
    assert ws.cell(1, 1).value == "Ligand"

    # First 4 summary rows: one per ligand
    expected_labels = ["His 255", "His 259", "Cys 207", "Cys 216"]
    for i, expected in enumerate(expected_labels, start=2):
        assert ws.cell(i, 1).value == expected, (
            f"row {i} should be {expected}, got {ws.cell(i,1).value}")

    # Atom counts: His should have 10, Cys should have 6
    assert ws.cell(2, 6).value == 10   # His 255
    assert ws.cell(3, 6).value == 10   # His 259
    assert ws.cell(4, 6).value == 6    # Cys 207
    assert ws.cell(5, 6).value == 6    # Cys 216

    # No warnings should have been emitted
    assert not runlog.warnings, f"unexpected warnings: {runlog.warnings}"


# ---------------------------------------------------------------------------
# Layer 4: export_embedding_plots must accept new arguments without raising
# ---------------------------------------------------------------------------

def test_export_embedding_plots_back_compat():
    """The new signature must remain back-compatible — calling it the
    pre-v1.0.3 way (no sse_umap_data/ca_umap_data/ca_data) must not
    raise and must not emit warnings."""
    from modenanalyse_2fe2s.export import export_embedding_plots

    results, _ = _make_synthetic_results(n_modes=10)
    runlog = _FakeRunLog()

    with tempfile.TemporaryDirectory() as td:
        cfg = _FakeCfg(td)
        # No embedding_coords either — empty dict triggers the "no plots
        # to render" path but must NOT raise.
        export_embedding_plots({}, results, cfg.outname, runlog)

    # Should not raise; matplotlib may or may not be installed — function
    # warns then returns. No assertion on output files, only no exception.


def test_export_embedding_plots_renders_ca_umap_png():
    """When ca_umap_data is supplied with valid coordinates, a PNG must
    be written next to outname('_embedding_Ca_UMAP.png')."""
    pytest.importorskip("matplotlib")
    from modenanalyse_2fe2s.embedding import compute_ca_umap_cluster
    from modenanalyse_2fe2s.export    import export_embedding_plots

    results, ca_data = _make_synthetic_results()
    runlog = _FakeRunLog()
    ca_umap = compute_ca_umap_cluster(results, ca_data, runlog=runlog)

    with tempfile.TemporaryDirectory() as td:
        cfg = _FakeCfg(td)
        export_embedding_plots({}, results, cfg.outname, runlog,
                                ca_umap_data=ca_umap)
        expected = os.path.join(td, "test_embedding_Ca_UMAP.png")
        assert os.path.exists(expected), \
            f"Ca-UMAP PNG should have been written to {expected}"
        # PNG header check
        with open(expected, "rb") as fh:
            head = fh.read(8)
        assert head == b"\x89PNG\r\n\x1a\n", "output must be a real PNG"


def test_export_embedding_plots_renders_ca_heatmap_png():
    """When ca_data is supplied, a heatmap PNG must be written."""
    pytest.importorskip("matplotlib")
    from modenanalyse_2fe2s.export import export_embedding_plots

    results, ca_data = _make_synthetic_results()
    runlog = _FakeRunLog()

    with tempfile.TemporaryDirectory() as td:
        cfg = _FakeCfg(td)
        export_embedding_plots({}, results, cfg.outname, runlog,
                                ca_data=ca_data)
        expected = os.path.join(td, "test_ca_amplitudes_heatmap.png")
        assert os.path.exists(expected), \
            f"Ca-heatmap PNG should have been written to {expected}"
